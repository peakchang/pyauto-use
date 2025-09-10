# func.py
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.alert import Alert
from openpyxl import load_workbook
import pyautogui as pg
import os
import random
import clipboard as cb
import ctypes
import pygetwindow as gw
from pywinauto import Desktop
from datetime import datetime, timezone
from zoneinfo import ZoneInfo  # Python 3.9+



def wait_ready(driver, timeout=15):
    """
    문서 readyState가 'complete'가 될 때까지 대기.
    """
    end = time.time() + timeout
    while time.time() < end:
        try:
            if driver.execute_script("return document.readyState") == "complete":
                return True
        except Exception:
            pass
        time.sleep(0.1)
    return False

def wait_visible(driver, locator, timeout=10):
    """
    특정 요소가 보일 때까지 대기.
    locator 예: (By.CSS_SELECTOR, "input[name='q']")
    """
    return WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located(locator)
    )

def band_join_loop(driver, cancel_event, stop_event):
    print('밴드 가입 들어옴!!!!')

    bandWb = load_workbook('./etc/band_work.xlsx')
    bandEx = bandWb.active

    try:
        with open(f'./etc/profile.txt', 'rt', encoding='UTF8') as f:
            profile = f.read()
    except:
        pass

    try:
        with open(f'./etc/profile.txt', 'r') as f:
            profile = f.read()
    except:
        pass

    if profile == '':
        pg.alert('프로필이 비어있습니다. profile.txt 파일을 확인해주세요.')
        return

    bandCount = 1
    onedayLimitChk = False
    while not cancel_event.is_set() and not stop_event.is_set():

        bandCount += 1
        if bandEx.cell(bandCount,1).value is None:
            break

        if onedayLimitChk:
            print('오늘 가입 가능한 밴드 수 초과!! 프로그램 종료합니다.')
            break

        if bandEx.cell(bandCount,3).value is not None:
            print('이미 가입된 밴드입니다. 다음 밴드로 넘어갑니다.')
            continue

        print(bandEx.cell(bandCount,1).value)

        url = bandEx.cell(bandCount,1).value
        driver.get(url)
        profileSuccessStatus = False
        

        # 밴드 가입 메인!! 가입하기 버튼이 사라지면 탈출!!
        while True:

            # 글쓰기 버튼이 있으면 이미 가입된 밴드!!
            try:
                wait_float(0.5, 1.0)
                writeBtn = driver.find_elements(by=By.CSS_SELECTOR, value="._btnPostWrite")
                if len(writeBtn) > 0:
                    break
            except:
                pass


            # 가입 버튼 클릭 > 가입 모달창 뜨면 break
            while True:
                print('가입 버튼 클릭 gogo')
                try:
                    wait_float(0.5, 1.0)
                    joinBtn = driver.find_element(by=By.CSS_SELECTOR, value="._btnJoinBand")
                    joinBtn.click()
                except Exception as e:
                    print(str(e))
                    pass

                
                # 모달 떠있으면 break!!
                try:
                    wait_float(0.5, 1.0)
                    joinModalChk = driver.find_element(by=By.CSS_SELECTOR, value=".lyContent")
                    headingMsg = joinModalChk.find_element(by=By.CSS_SELECTOR, value=".headingMsg")
                    if '1일 가입' in headingMsg.text:
                        print('오늘 가입 가능한 밴드 수 초과!!')
                        onedayLimitChk = True
                        break
                    elif joinModalChk:
                        break
                except Exception as e:
                    print(str(e))
                    break

            if onedayLimitChk == True:
                break
            # 프로필 여러개일 시 프로필 리스트 가져오기!!
            while True:
                try:
                    wait_float(1.2, 1.5)
                    profileList = driver.find_elements(by=By.CSS_SELECTOR, value=".textEllipsis")
                    if len(profileList) > 0:
                        break
                except Exception as e:
                    print(str(e))
                    pass

            # 프로필 이름 대조 > 있으면 radio input 클릭!!
            for i, p in enumerate(profileList):
                if p.text == profile:

                    while True:
                        profileRadioList = driver.find_elements(by=By.CSS_SELECTOR, value=".checkInput._radio")
                        profileRadioList[i].click()
                        break
                    profileSuccessStatus = True
                    break

            if profileSuccessStatus == False:
                while True:

                    # 새 프로필 버튼 클릭
                    try:
                        wait_float(0.5, 1.0)
                        newProfileBtn = driver.find_element(by=By.CSS_SELECTOR, value="._newProfileBtn")
                        newProfileBtn.click()
                    except:
                        pass

                    # 프로필 INPUT 창 체크 후 프로필 복붙 하기!
                    try:
                        focus_window('Chrome')
                        wait_float(0.5, 1.0)
                        profileNameInputBtn = driver.find_element(by=By.CSS_SELECTOR, value=".uInput")
                        if profileNameInputBtn:
                            cb.copy(profile)
                            wait_float(0.5,1.2)
                            profileNameInputBtn.click()
                            wait_float(0.5,1.2)
                            pg.hotkey('ctrl', 'v')
                            wait_float(0.5, 1.2)
                            profileNameInputBtn.send_keys(profile)
                    except:
                        pass
                    try:
                        profileNameInputBtnChk = driver.find_element(by=By.CSS_SELECTOR, value="._nameInput")
                        if profileNameInputBtnChk.get_attribute('value') == profile:
                            break
                        else:
                            continue
                    except:
                        continue

                
            # 가입하기 버튼 클릭하기! (새 프로필 생성할때는 가입하기 버튼이 두개임!!)
            while True:
                try:
                    confirmSuccessStatus = False
                    wait_float(0.5, 1.0)
                    confirmBtns = driver.find_elements(by=By.CSS_SELECTOR, value="._confirmBtn")
                    for confirmBtn in confirmBtns:
                        try:
                            confirmBtn.click()
                            confirmSuccessStatus = True
                        except:
                            pass
                    if confirmSuccessStatus:
                        break
                except Exception as e:
                    continue

            # alert 창 뜨면 확인 버튼 클릭
            try:
                wait_float(0.5, 1.0)
                alert = driver.switch_to.alert # type: Alert
                print("알림 내용:", alert.text)
                alert.accept() # 또는 alert.dismiss()
            except:
                pass

        
        bandEx.cell(bandCount,3).value = 'ok'
        bandWb.save('./etc/band_work.xlsx')
    
                        

                    




def band_write_loop(driver, cancel_event, stop_event,):
    
    bandWb = load_workbook('./etc/band_work.xlsx')
    bandEx = bandWb.active

    bandCount = 1
    while not cancel_event.is_set() and not stop_event.is_set():

        focus_window('Chrome')

        postWriteBool = True

        bandCount += 1

        if bandEx.cell(bandCount,4).value is not None:
            continue

        if bandEx.cell(bandCount,4).value is None and bandEx.cell(bandCount,3).value is None:
            for i in range(bandCount):
                if i == 0:
                    continue
                bandEx.cell(i + 1 , 4).value = None
                bandWb.save('./etc/band_work.xlsx')
            bandCount = 1
            
            continue

        

        url = bandEx.cell(bandCount,1).value
        driver.get(url)

        focus_window('Chrome')



        
        while True:
            focus_window('Chrome')
            try:
                wait_float(0.8,1.5)
                latestChk = driver.find_element(by=By.CSS_SELECTOR, value=".postSet._btnPostMore")
                latestChk.click()
            except Exception as e:
                print(str(e))
                continue

            try:
                wait_float(0.8,1.5)
                firstMenu = driver.find_element(by=By.CSS_SELECTOR, value=".lyMenu._postMoreMenu")
                menuList = firstMenu.find_elements(by=By.CSS_SELECTOR, value="._postMoreMenuUl li")

                try:
                    print(len(menuList))
                    print(menuList[3].text)
                    print(menuList[4].text)
                    print(menuList[5].text)
                    if '삭제하기' in menuList[5].text:
                        postWriteBool = False
                except:
                    pass

                break
            except Exception as e:
                print(str(e))
                continue
        
        if postWriteBool == False:
            print('이미 게시글이 존재합니다. 다음 밴드로 넘어갑니다.')
            bandEx.cell(bandCount,4).value = 'skip'
            bandWb.save('./etc/band_work.xlsx')
            continue


            

        


        # 멤버 _lnbMenus[4] > 내 정보 _btnSetting > 내 게시글 _btnGotoSearchMemberContent


        # 멤버 > 내 게시글 > 오래된순 정렬까지!!
        while True:
            focus_window('Chrome')
            try:
                wait_float(0.8,1.5)
                memberBtn = driver.find_elements(by=By.CSS_SELECTOR, value=".lnbTopMenuItemLink")
                memberBtn[4].click()
            except Exception as e:
                print(str(e))
                pass

            try:
                wait_float(0.8,1.5)
                settingBtn = driver.find_element(by=By.CSS_SELECTOR, value="._btnSetting")
                settingBtn.click()
            except Exception as e:
                print(str(e))
                pass

            try:
                wait_float(0.8,1.5)
                myPostListBtn = driver.find_element(by=By.CSS_SELECTOR, value="._btnGotoSearchMemberContent")
                myPostListBtn.click()
            except Exception as e:
                print(str(e))
                pass

            try:
                wait_float(0.8,1.5)
                sortSelectBtn = driver.find_element(by=By.CSS_SELECTOR, value=".buttonSorting._btnSort")
                sortSelectBtn.click()
            except Exception as e:
                print(str(e))
                pass

            try:
                wait_float(1.2,1.9)
                selectList = driver.find_elements(by=By.CSS_SELECTOR, value="._optionMenuLink")
                print(selectList)
                selectList[1].click()
                break
            except Exception as e:
                print(str(e))
                pass


        # 게시글 삭제 루프!! 최신 1개 남기고 다 지우기!!
        while True:
            focus_window('Chrome')
            try:
                wait_float(0.8,1.5)
                myPostList = driver.find_elements(by=By.CSS_SELECTOR, value=".postSet._btnPostMore")
                print(myPostList)
                print(len(myPostList))
                if len(myPostList) <= 1:
                    break
                else:
                    myPostList[0].click()
            except Exception as e:
                print('메뉴 버튼 wrap 오픈 에러')
                print(str(e))
                pass

            try:
                wait_float(0.8,1.5)
                myPostMenu = driver.find_element(by=By.CSS_SELECTOR, value=".lyMenu._postMoreMenu")
                deleteBtns = myPostMenu.find_elements(by=By.CSS_SELECTOR, value="._postMoreMenuUl li")
                deleteBtns[5].click()
            except Exception as e:
                print('메뉴 버튼 클릭 에러')
                print(str(e))
                pass

            try:
                wait_float(0.8,1.5)
                confirmBtn = driver.find_element(by=By.CSS_SELECTOR, value="._btnConfirm")
                confirmBtn.click()
            except Exception as e:
                print('삭제 확인 버튼 클릭 에러')
                print(str(e))
                pass
        

        # 글쓰기 시자아아아아아아아아아악~~~~~~~~~~~~~~~~~~~~~~~~~~~~~


        # 이미지 업로드 하기!!!!!!
        contentSuccessStatus = False
        errCount = 0
        while True:
            errCount += 1
            if errCount > 10:
                pg.press('F5')
                errCount = 0
            try:
                wait_float(0.5,1.2)
                postWriteBtn = driver.find_element(by=By.CSS_SELECTOR, value="._btnPostWrite")
                postWriteBtn.click()
            except:
                pass

            try:
                wait_float(0.5,1.2)
                postWriteModal = driver.find_elements(by=By.CSS_SELECTOR, value=".lyWrap._lyWrap")
                if len(postWriteModal) == 0:
                    continue
            except:
                pass

            if contentSuccessStatus == False:
                try:
                    wait_float(0.5,1.2)
                    imgUploadBtns = driver.find_elements(by=By.CSS_SELECTOR, value="._btnAttachPhoto")
                    print(f"imgUploadBtns 갯수 : {len(imgUploadBtns)}")
                    for imgUploadBtn in imgUploadBtns:
                        imgUploadBtn.click()
                except Exception as e:
                    print(str(e))
                    print('파일 업로드 버튼 클릭 에러!!')
                    pass

                try:
                    wait_float(0.5,1.2)
                    if is_window_open('열기'):
                        print("파일 선택 창이 이미 열려 있습니다.")
                    else:
                        continue

                    focus_window('열기')

                    print("파일 선택 창이 열려 있습니다.")

                    nowPath = os.getcwd()
                    content_image_path = nowPath + f"\etc\content"

                    cb.copy(content_image_path)
                    wait_float(0.5,1.2)
                    pg.hotkey('ctrl', 'v')
                    wait_float(0.5, 1.2)
                    pg.press('enter')
                    wait_float(1.5, 2.5)


                    content_image_path = nowPath + f"\etc\content\image"

                    cb.copy(content_image_path)
                    wait_float(0.5,1.2)
                    pg.hotkey('ctrl', 'v')
                    wait_float(0.5, 1.2)
                    pg.press('enter')
                    wait_float(1.2, 1.5)


                except Exception as e:
                    print(str(e))
                    pass



                try:
                    wait_float(0.5,1.2)
                    focus_window('Chrome')
                    imgSubmitBtn = driver.find_element(by=By.CSS_SELECTOR, value=".uButton.-confirm._submitBtn")
                    imgSubmitBtn.click()
                except Exception as e:
                    print('파일 업로드 완료 버튼 클릭 에러!!')
                    pg.press('esc')
                    pg.press('esc')
                    continue


                content = ""
                while True:

                    try:
                        with open(f'./etc/content/content.txt', 'rt', encoding='UTF8') as f:
                            content = f.read()
                    except:
                        pass

                    try:
                        with open(f'./etc/content/content.txt', 'r') as f:
                            content = f.read()
                    except:
                        pass

                    if content == '':
                        continue
                    else:
                        break

                try:
                    focus_window('Chrome')
                    pg.press('enter')
                    pg.press('enter')
                    pg.press('enter')
                    cb.copy(content)
                    wait_float(0.5,1.2)
                    pg.hotkey('ctrl', 'v')
                    wait_float(0.5, 1.2)
                    pg.press('enter')
                    wait_float(1.2, 1.5)
                except:
                    pass

            try:
                wait_float(0.5,1.2)
                contentChk = driver.find_elements(by=By.CSS_SELECTOR, value=".cke_widget_wrapper.cke_widget_block")
                print(f"contentChk 갯수 : {len(contentChk)}")
                if len(contentChk) == 0:
                    pg.press('F5')
                    wait_float(1.5,2.5)
                    continue
                else:
                    contentSuccessStatus = True

                submitBtn = driver.find_element(by=By.CSS_SELECTOR, value="._btnSubmitPost")
                submitBtn.click()
                break
                 
            except:
                pass


        

        try:
            wait_float(2.5,3.5)
            btnConfirm = driver.find_element(by=By.CSS_SELECTOR, value=".uButton.-confirm._btnConfirm")
            btnConfirm.click()
        except:
            pass
            

        bandEx.cell(bandCount,4).value = 'ok'
        bandWb.save('./etc/band_work.xlsx')

        wait_float(20.0, 30.0)
        
            





        pass





# -----------------------------------------------------

def wait_float(start, end):
    wait_ran = random.uniform(start, end)
    time.sleep(wait_ran)



def focus_window(winName):
    while True:
        try:
            user32 = ctypes.windll.user32
            foreground_window = user32.GetForegroundWindow()
            window = gw.Window(foreground_window)
            print(window.title)
            if winName in window.title:
                break
            else:
                windows = Desktop(backend="uia").windows()
                for window in windows:
                    if winName in window.window_text():
                        window.set_focus()
                        break
        except Exception as e:
            print(str(e))
            pass


def is_window_open(title_substr: str) -> bool:
    return any(title_substr.lower() in (t or "").lower() for t in gw.getAllTitles())

