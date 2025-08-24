import random
import threading
import string
import time
from datetime import datetime, timedelta
import sys
import os
from pathlib import Path
import json
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import WebDriverException
# from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
import pyautogui as pg
from openpyxl import load_workbook
from tkinter import *
from tkinter import ttk
import winsound as ws
import winsound as sd
import shutil
import getpass
import os
import clipboard as cb
import requests
from PIL import Image
from io import BytesIO


def goScript(getDict):
    if getDict['work_site'] == "":
        pg.alert('작업 사이트를 선택 하세요')
        sys.exit(0)

    try:
        pcUser = getpass.getuser()
        options = Options()
        # user_data = f'C:\\Users\\{pcUser}\\AppData\\Local\\Google\\Chrome\\User Data\\default'
        # options.add_argument(f"user-data-dir={user_data}")
        # options.add_argument(f'--profile-directory=Profile {profileInfo['pl_number']}')
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        driver = webdriver.Chrome(options=options)

        if getDict['work_site'] == 'naver':
            driver.get('https://naver.com')
        else:
            driver.get('https://google.com')
        driver.set_page_load_timeout(12)
        driver.set_window_size(1300, 800)
        driver.set_window_position(0,0)
    except Exception as e:
        print(e)
        print('크롬 창 오픈 실패!!')
        if driver:
            driver.quit()
        pass

    imgWb = load_workbook('./etc/img_work.xlsx')
    imgEx = imgWb.active

    exCountTemp = 0
    while True:
        exCountTemp += 1
        if imgEx.cell(exCountTemp,4).value is None:
            break

    
    pg.alert(f'{imgEx.cell(exCountTemp, 3).value} 부터 시작합니다~')
    exCount = exCountTemp - 1
    
    
    while True:
        exCount += 1
        if imgEx.cell(exCount, 3).value is None or imgEx.cell(exCount, 1).value is None:
            break


        folderPath = f'./etc/img/{imgEx.cell(exCount, 1).value}'

        print(imgEx.cell(exCount, 3).value)
        cb.copy(imgEx.cell(exCount, 3).value)
        if getDict['work_site'] == 'naver':
            while True:
                try:
                    searchBar = driver.find_element(by=By.CSS_SELECTOR, value='#query')
                    searchBar.click()
                    time.sleep(0.5)
                    pg.hotkey('ctrl','a')
                    time.sleep(0.5)
                    pg.hotkey('ctrl','v')
                except:
                    pass

                try:
                    searchBar = driver.find_element(by=By.CSS_SELECTOR, value='.box_window')
                    searchBar.click()
                    time.sleep(0.5)
                    pg.hotkey('ctrl','a')
                    time.sleep(0.5)
                    pg.hotkey('ctrl','v')
                except:
                    pass

                try:
                    searchBarChk = driver.find_element(by=By.CSS_SELECTOR, value='#query')
                    print(searchBarChk.get_attribute('value'))
                    if searchBarChk.get_attribute('value')!= '' and searchBarChk.get_attribute('value') is not None:
                        time.sleep(0.5)
                        pg.press('enter')
                        break
                except:
                    pass

                try:
                    searchBarChk = driver.find_element(by=By.CSS_SELECTOR, value='.box_window')
                    print(searchBarChk)
                    print(searchBarChk.get_attribute('value'))
                    if searchBarChk.get_attribute('value')!= '' and searchBarChk.get_attribute('value') is not None:
                        time.sleep(0.5)
                        pg.press('enter')
                        break
                except:
                    pass

        else:
            # 구글 검색~~
            while True:
                
                try:
                    searchBar = driver.find_element(by=By.CSS_SELECTOR, value='textarea.gLFyf')
                    searchBar.click()
                    time.sleep(0.5)
                    pg.hotkey('ctrl','a')
                    time.sleep(0.5)
                    pg.hotkey('ctrl','v')
                except Exception as e:
                    print(str(e))
                    pass

                try:
                    searchBarChk = driver.find_element(by=By.CSS_SELECTOR, value='textarea.gLFyf')
                    print(searchBarChk.get_attribute('value'))
                    if searchBarChk.get_attribute('value')!= '' and searchBarChk.get_attribute('value') is not None:
                        time.sleep(0.5)
                        pg.press('enter')
                        break
                except:
                    pass
        
        while True:
            pg.alert('이미지 다운받을 이미지를 새 탭에 띄워주세요!!')
            window_handles = driver.window_handles
            if len(window_handles) == 1:
                pg.alert('다운받을 이미지 창이 뜨지 않았습니다 다시!')
                continue

            ensure_folder_exists(f'./etc/img/{imgEx.cell(exCount, 1).value}')
            

            # 첫번째 이미지 엑셀에 입력
            window_handles = driver.window_handles
            driver.switch_to.window(window_handles[2])
            print(f"3번째 창의 주소: {driver.current_url}")
            savePath = f'{folderPath}/{generate_random_string()}.jpg'
            returnStatus = download_and_convert_to_jpg(driver.current_url, savePath)
            driver.close()  # 3번째 창 닫기

            if returnStatus == False:
                window_handles = driver.window_handles
                driver.switch_to.window(window_handles[0])
                delete_files_in_folder(folderPath)
                pg.alert('이미지 다운 오류 다시 시도해주세요!')
                continue

            

            # 두번째 이미지 엑셀에 입력
            window_handles = driver.window_handles
            driver.switch_to.window(window_handles[1])
            print(f"3번째 창의 주소: {driver.current_url}")
            savePath = f'{folderPath}/{generate_random_string()}.jpg'
            returnStatus = download_and_convert_to_jpg(driver.current_url, savePath)
            driver.close()  # 3번째 창 닫기

            
            if returnStatus == False:
                window_handles = driver.window_handles
                driver.switch_to.window(window_handles[0])
                delete_files_in_folder(folderPath)
                pg.alert('이미지 다운 오류 다시 시도해주세요!')
                continue

            window_handles = driver.window_handles
            driver.switch_to.window(window_handles[0])

            break
        imgEx.cell(exCount, 4).value = 'OK'
        imgWb.save('./etc/img_work.xlsx')
    
    pg.alert('종료합니다.')
    sys.exit(0)


def get_nidx_list(getDict):

    siteLink = "https://happy-toad2.shop"
    # siteLink = "http://localhost:3020"
    if getDict['start_val'] is None or getDict['start_val'] == '':
        pg.alert('시작 값이 없습니다. 종료합니다!')
        sys.exit(0)
        
    while True:
        try:
            res = requests.post(f"{siteLink}/api/v7/res_blog/get_idx_list",{'start_val' : getDict['start_val'], 'count_val' : getDict['count_val']}).json()
            if res['status'] == True:
                print(res)
                result = "\n".join(
                    f"{item['n_idx']}\t{'OK' if item['n_link_use'] == 1 else ''}"
                    for item in res['idx_list']
                )

                cb.copy(result)
                break
        except Exception as e:
            print(str(e))
            pass

    pg.alert('완룡~')



def check_alive_blog(getDict):
    nowNum = int(getDict['start_val'])
    siteLink = "https://happy-toad2.shop"
    # siteLink = "http://localhost:3020"

    while True:
        try:
            pcUser = getpass.getuser()
            options = Options()
            # user_data = f'C:\\Users\\{pcUser}\\AppData\\Local\\Google\\Chrome\\User Data\\default'
            # options.add_argument(f"user-data-dir={user_data}")
            # options.add_argument(f'--profile-directory=Profile {profileInfo['pl_number']}')
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            driver = webdriver.Chrome(options=options)
            driver.get(f'https://www.google.com')
            driver.set_page_load_timeout(12)
            driver.set_window_size(1300, 800)
            driver.set_window_position(0,0)
            break
        except Exception as e:
            print(e)
            print('크롬 창 오픈 실패!!')
            if driver:
                driver.quit()
            pass
    while True:
        while True:
            try:
                res = requests.post(f"{siteLink}/api/v7/res_blog/get_chk_blog_id_info",{'start_val' : nowNum}).json()
                if res['status'] == True:
                    if 'OK' in res['id_info']['n_memo2']:
                        nowNum += 1
                        time.sleep(0.3)
                        continue
                    elif '작업' not in res['id_info']['n_memo2']:
                        pg.alert('작업 안됨! 종료합니다!!')
                        driver.quit()
                        sys.exit(0)

                    else:
                        print(res['id_info'])
                        driver.get(f'https://blog.naver.com/{res['id_info']['n_id']}')
                        break
            except Exception as e:
                print(str(e))
                pass
        

        while True:
            try:
                driver.switch_to.frame('mainFrame')
                searchTitle = driver.find_elements(by=By.CSS_SELECTOR, value='.se-title-text')
                print(searchTitle)
                if len(searchTitle) > 0:
                    cb.copy(searchTitle[0].text)
                    driver.switch_to.default_content()
                    break
            except Exception as e:
                print(str(e))
                pass

        driver.get(f'https://naver.com')
        while True:
            try:
                searchBar = driver.find_element(by=By.CSS_SELECTOR, value='#query')
                searchBar.click()
                time.sleep(0.5)
                pg.hotkey('ctrl','a')
                time.sleep(0.3)
                pg.hotkey('ctrl','v')
                time.sleep(0.5)
                pg.press('enter')
                searchSuccessChk = driver.find_elements(by=By.CSS_SELECTOR, value='.lnb_group')
                if len(searchSuccessChk) > 0:
                    break                
            except:
                pass

        while True:
            try:
                boxList1 = driver.find_elements(by=By.CSS_SELECTOR, value='.source_box')
                boxList2 = driver.find_elements(by=By.CSS_SELECTOR, value='.user_box_inner')
                boxList = boxList1 + boxList2
                if len(boxList) > 0:
                    break
            except:
                pass
        
        searchResStatus = False
        for sourceBox in boxList:
            try:
                getLink = sourceBox.find_element(by=By.CSS_SELECTOR, value='a').get_attribute('href')
                print(getLink)
                if res['id_info']['n_id'] in getLink:
                    print('성공 아니야?!?!?!')
                    searchResStatus = True
                    break
            except:
                pass
        
        while True:
            try:
                res = requests.post(f"{siteLink}/api/v7/res_blog/update_chk_blog",{'search_status' : searchResStatus, 'n_id' : res['id_info']['n_id'], 'n_memo2' : res['id_info']['n_memo2']}).json()
                print(res)
                if res['status'] == True:
                    break
            except Exception as e:
                print(str(e))
                pass

        nowNum = int(nowNum) + 1

def download_and_convert_to_jpg(image_url, save_path):
    try:
        # 이미지 다운로드
        response = requests.get(image_url, stream=True)
        response.raise_for_status()  # HTTP 오류 확인

        # 이미지를 메모리에 로드
        image = Image.open(BytesIO(response.content))

        # JPG로 변환 (RGB 모드로 변환)
        if image.mode != "RGB":
            image = image.convert("RGB")

        # 저장
        image.save(save_path, "JPEG")
        print(f"이미지가 성공적으로 저장되었습니다: {save_path}")
        return True
    except Exception as e:
        print(f"오류 발생: {e}")
        return False
        

def ensure_folder_exists(folder_path):
    """
    폴더가 존재하지 않으면 생성합니다.
    """
    if not os.path.exists(folder_path):  # 폴더 존재 여부 확인
        os.makedirs(folder_path)        # 폴더 생성
        print(f"폴더를 생성했습니다: {folder_path}")
    else:
        print(f"폴더가 이미 존재합니다: {folder_path}")

def generate_random_string(length=7):
    """
    지정된 길이의 영어 소문자만 포함된 랜덤 문자열을 생성합니다.
    """
    characters = string.ascii_lowercase  # 영어 소문자만 포함
    random_string = ''.join(random.choices(characters, k=length))
    return random_string


def delete_files_in_folder(folder_path):
    """
    지정된 폴더 내 모든 파일을 삭제합니다.
    
    Parameters:
        folder_path (str): 파일을 삭제할 폴더 경로
    """
    if not os.path.exists(folder_path):
        print(f"Error: Folder '{folder_path}' does not exist.")
        return
    
    for file_name in os.listdir(folder_path):
        file_path = os.path.join(folder_path, file_name)
        try:
            if os.path.isfile(file_path):  # 파일인지 확인
                os.remove(file_path)
                print(f"Deleted: {file_path}")
        except Exception as e:
            print(f"Error deleting {file_path}: {e}")


def get_latest_link(getDict):

    siteLink = "https://happy-toad2.shop"
    # siteLink = "http://localhost:3020"

    # siteLinks = [{"site":"https://atb.co.kr/","board" : "view"}, {"site":"https://tssp.co.kr/","board" : "view_board"}]

    siteLinks = [{"site":"https://atb.co.kr/", "board":"view", "version":"v3"}, {"site":"https://tssp.co.kr/","board" : "view_board", "version":"v3"}]

    for site in siteLinks:
        
        while True:
            try:
                time.sleep(2.5)
                res = requests.post(f"{site['site']}/api/{site['version']}/get_latest_list",{'link_count' : getDict['link_count'],'board' : site['board']}).json()
                if res['status'] == True:
                    print(res)
                    result = "\n".join(
                        f"{item['bo_subject'] if item['bo_subject'] else item['bo_name']}\t{site["site"]}view/{item['bo_id']}"
                        for item in res['latest_list']
                    )
                    cb.copy(result)
                    break
            except Exception as e:
                print(str(e))
                pass

        pg.alert(f"{site["site"]} 복사 완료!! 붙여넣기 하세요!!")

    pg.alert('완료우!!!')