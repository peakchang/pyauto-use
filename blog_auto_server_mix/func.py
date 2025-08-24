import random
import threading
import time
from datetime import datetime, timedelta
import sys
import os
from pathlib import Path
# from typing import Optional
# from pyparsing import And
import requests
# from bs4 import BeautifulSoup as bs
import json
import re
import pyautogui as pg
import pyperclip
import pygetwindow as gw
import ctypes
from pywinauto import Desktop

import clipboard as cb
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import WebDriverException
from selenium.common.exceptions import TimeoutException, NoSuchElementException

from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from ppadb.client import Client as AdbClient
import keyboard
from tkinter import *
from tkinter import ttk
import winsound as ws
import winsound as sd
import shutil
import getpass
from PIL import Image
import os
from openai import OpenAI



def naverIdChkFuncOnlyName(driver,getDict,blogInfo,siteLink,workBlogNum):
    while True:
        try:
            wait_float(1.2,1.9)
            aside = driver.find_element(by=By.CSS_SELECTOR, value=".sha_ico_aside")
            aside.click()
        except:
            pass

        try:
            wait_float(1.2,1.9)
            linkUser = driver.find_element(by=By.CSS_SELECTOR, value=".link_user")
            linkUser.click()
        except:
            pass

        try:
            wait_float(1.2,1.9)
            userInfoPageChk = driver.find_element(by=By.CSS_SELECTOR, value=".row_item.name")
            if userInfoPageChk:
                break
        except:
            pass
    
    while True:
        # 메모에 이름이 정상적으로 있는지 체크하기!!!
        try:
            wait_float(1.2,1.9)
            userName = driver.find_element(by=By.CSS_SELECTOR, value=".row_item.name .item_text")
            print(userName.text)
            if userName.text not in blogInfo['n_memo1']:
                loginChkStatus = False
                for i in range(3):
                    fr = 1550    # range : 37 ~ 32767
                    du = 300     # 1000 ms ==1second
                    sd.Beep(fr, du)
                pg.alert(f'{blogInfo['n_id']} 메모 내 이름이 없습니다.')
                while True:
                    try:
                        res = requests.get(f"{siteLink}/api/v7/res_blog/get_blog_id_info_m?get_profile={workBlogNum}").json()
                        if res['status'] and res['blog_info']:
                            blogInfo = res['blog_info']
                            break
                    except Exception as e:
                        print('에러요~~~')
                        print(str(e))
                        pass
                    wait_float(2.5,3.5)
            else:
                return

        except Exception as e:
            print(str(e))
            pass
    
def naverIdChkFunc(driver,getDict,blogInfo,siteLink,workBlogNum):
    # 여기서 이메일 인증 / 이름 맞는지 체크체크!!!!
    while True:
        try:
            wait_float(1.2,1.9)
            aside = driver.find_element(by=By.CSS_SELECTOR, value=".sha_ico_aside")
            aside.click()
        except:
            pass

        try:
            wait_float(1.2,1.9)
            linkUser = driver.find_element(by=By.CSS_SELECTOR, value=".link_user")
            linkUser.click()
        except:
            pass

        try:
            wait_float(1.2,1.9)
            userInfoPageChk = driver.find_element(by=By.CSS_SELECTOR, value=".row_item.name")
            if userInfoPageChk:
                break
        except:
            pass

    if getDict['directVal'] == 'plus':
        chkBlogNum = blogInfo['n_blog_order'] + 200
    elif getDict['directVal'] == 'minus':
        chkBlogNum = blogInfo['n_blog_order'] - 200

    while True:
        # 메모에 이름이 정상적으로 있는지 체크하기!!!
        try:
            wait_float(1.2,1.9)
            userName = driver.find_element(by=By.CSS_SELECTOR, value=".row_item.name .item_text")
            print(userName.text)
            if userName.text not in blogInfo['n_memo1']:
                loginChkStatus = False
                for i in range(3):
                    fr = 1550    # range : 37 ~ 32767
                    du = 300     # 1000 ms ==1second
                    sd.Beep(fr, du)
                pg.alert(f'{blogInfo['n_id']} 메모 내 이름이 없습니다.')
                while True:
                    try:
                        res = requests.get(f"{siteLink}/api/v7/res_blog/get_blog_id_info_m?get_profile={workBlogNum}").json()
                        if res['status'] and res['blog_info']:
                            blogInfo = res['blog_info']
                            break
                    except Exception as e:
                        print('에러요~~~')
                        print(str(e))
                        pass
                    wait_float(2.5,3.5)
            else:
                loginChkStatus = True

        except Exception as e:
            print(str(e))
            pass

        # 이메일 인증 체크할 아이디 구하기!!!
        while True:
            try:
                res = requests.get(f"{siteLink}/api/v7/res_blog/get_blog_id_info_m_profile?get_profile={chkBlogNum}").json()
                if res['status'] and res['chk_blog_info']:
                    chkBlogInfo = res['chk_blog_info']
                    break
            except Exception as e:
                pass
            wait_float(2.5,3.5)

        # 이메일 인증이 아예 없을경우 continue!!!
        try:
            emailRegArea = driver.find_element(by=By.CSS_SELECTOR, value="#pswdEmailRegDiv #pswdEmailRegSpan")
            if '없음' in emailRegArea.text:
                loginChkStatus = False
                for i in range(3):
                    fr = 1550    # range : 37 ~ 32767
                    du = 300     # 1000 ms ==1second
                    sd.Beep(fr, du)
                pg.alert(f'{blogInfo['n_id']} 이메일 주소가 존재하지 않습니다.')
                continue
        except:
            pass
        # 이메일 인증이 맞게 되어 있는지 확인하기!!

        try:
            emailRegBtn = driver.find_element(by=By.CSS_SELECTOR, value="#pswdEmailRegBtn")
            emailRegBtn.click()
            wait_float(1.5,2.2)

            emailInput = driver.find_element(by=By.CSS_SELECTOR, value="#confirmPswdEmail")
            emailInput.send_keys(chkBlogInfo['n_id'] + '@naver.com')
            wait_float(1.5,2.2)

            emailChkBtn = driver.find_element(by=By.CSS_SELECTOR, value="#confirmPswdEmailDiv .btn_contact")
            emailChkBtn.click()
            wait_float(1.5,2.2)
            
            emailStatusEle = driver.find_element(by=By.CSS_SELECTOR, value="#e_pswdEmail")
            if '확인해 주세요' in emailStatusEle.text:
                loginChkStatus = False
                for i in range(3):
                    fr = 1550    # range : 37 ~ 32767
                    du = 300     # 1000 ms ==1second
                    sd.Beep(fr, du)
                pg.alert(f'{blogInfo['n_id']} 인증 이메일 주소를 확인해주세요!!')
            elif '확인되었습니다' in emailStatusEle.text:
                loginChkStatus = True
            
            closePopupBtn = driver.find_element(by=By.CSS_SELECTOR, value="#pswdEmailChangePopUpLayer .close_popup")
            closePopupBtn.click()
        except:
            pass

        if loginChkStatus == True:
            break

def writeBlogMobile(driver,workBlogNum,contentArr):
    pg.click(50,300)
    # 먼저 블로그 탭 들어가기!!!

    errCount = 0
    while True:
        errCount += 1
        if errCount > 10:
            errCount = 0
            driver.get('https://m.naver.com')
            wait_float(3.5,4.5)
        print('블로그 탭 진입~~~')
        try:
            wait_float(0.6,1.2)
            menuList = driver.find_elements(by=By.CSS_SELECTOR, value=".shs_item")
            for menu in menuList:
                if '블로그' in menu.text:
                    menu.click()
                    break
        except Exception as e:
            print(str(e))
            pass

        

        try:
            wait_float(0.6,1.2)
            blogPageChk = driver.find_elements(by=By.CSS_SELECTOR, value=".blog_btn__oKX22")
            if len(blogPageChk) > 0:
                break
        except:
            pass

        # 혹시 모를 모달 방짖ㄴ
        try:
            wait_float(0.6,1.2)
            closeQuickMenuBtn = driver.find_element(by=By.CSS_SELECTOR, value=".ah_link_landing.ah_close")
            closeQuickMenuBtn.click()
        except:
            pass

    # 우측 메뉴 열고 블로그 글쓰기 클릭!
    while True:
        # 혹시 모를 모달 방지
        try:
            wait_float(0.6,1.2)
            closeQuickMenuBtn = driver.find_element(by=By.CSS_SELECTOR, value=".ah_link_landing.ah_close")
            closeQuickMenuBtn.click()
        except:
            pass

        try:
            wait_float(0.6,1.2)
            closeModalBtn = driver.find_element(by=By.CSS_SELECTOR, value=".btn_text")
            closeModalBtn.click()
        except:
            pass
        try:
            wait_float(1.2,1.9)
            menuOpenBtn = driver.find_element(by=By.CSS_SELECTOR, value=".overflow_menu_btn__pdlpl")
            menuOpenBtn.click()
        except:
            pass

        try:
            wait_float(1.2,1.9)
            modalMenuList = driver.find_elements(by=By.CSS_SELECTOR, value=".profile_menu_item_link__zWetb")
            for menu in modalMenuList:
                if '글쓰기' in menu.text:
                    menu.click()
                    break
        except:
            pass

        try:
            wait_float(1.2,1.9)
            writePageChk = driver.find_elements(by=By.CSS_SELECTOR, value=".editor-header")
            if len(writePageChk) > 0:
                break
        except:
            pass
    
    # 제목 왼쪽 / 본문 가운데 정렬을 위한 작업
    contentAlignStatus = False
    alignArr = ['align-right','align-justify','align-left']
    while True:
        try:
            wait_float(0.5,1.2)
            editorArea = driver.find_element(by=By.CSS_SELECTOR, value=".se_editable")
            editorArea.click()
        except:
            pass
        try:
            wait_float(0.5,1.2)
            toolbarList = driver.find_element(by=By.CSS_SELECTOR, value=".toolbarList")
            toolbarList.click()
        except:
            pass
        for align in alignArr:
            try:
                wait_float(0.5,1.2)
                alignBtn = driver.find_element(by=By.CSS_SELECTOR, value=f".toolbarBtn.{align}")
                alignBtn.click()
                alignCenter = driver.find_elements(by=By.CSS_SELECTOR, value=".toolbarBtn.align-center")
                if len(alignCenter) > 0:
                    contentAlignStatus = True
                    break
                break
            except:
                pass
        
        if contentAlignStatus == False:
            continue
        else:
            break

    wait_float(1.2,1.9)


    lineCount = -1
    imageUploadCount = 0
    httpsEnter = 0
    while True:

        lineCount += 1

        if lineCount > len(contentArr):
            break
        
        focus_window('Chrome')

        # 리스트 범위 내에거 찾아야함!!
        try:
            getline = contentArr[lineCount]
        except:
            continue


        chkAction = getline.split('|')

        if 'http' in chkAction[0]:
            if httpsEnter == 0:
                pg.press('enter')
                pg.press('enter')
                pg.press('enter')
                pg.press('enter')
                pg.press('enter')
            httpsEnter += 1

            focus_window('Chrome')
            keyboard.write(text=getline, delay=0.12)
            wait_float(0.3,0.9)
            pg.press('enter')
            wait_float(1.5,1.9)

            errCount = 0
            deleteWork = False
            while True:
                wait_float(1.2,1.9)
                errCount += 1
                if errCount > 3:
                    break
                try:
                    delBtn = driver.find_element(by=By.CSS_SELECTOR, value=".toolbarBtn.remove")
                    delBtn.click()
                    deleteWork = True
                    break
                except:
                    pass

            while True:
                try:
                    wait_float(0.6, 1.2)
                    nextLineArr = driver.find_elements(by=By.CSS_SELECTOR, value=".se_link")
                    nextLineArr[-1].click()
                    break
                except Exception as e:
                    print(str(e))
                    pass
            for i in range(len(chkAction[0])):
                pg.press('right')
            if deleteWork:
                pg.press('enter')
            
            continue
        
        if chkAction[0] == 'img_line':
            imageUploadCount += 1
            nowPath = os.getcwd()

            errCount = 0
            while True:
                errCount += 1
                if errCount > 10:
                    while True:
                        try:
                            wait_float(1.2,1.9)
                            naverMainChk = driver.find_elements(by=By.CSS_SELECTOR, value="#MM_logo")
                            if len(naverMainChk) > 0:
                                break
                        except:
                            pass
                        driver.get('https://naver.com')
                        wait_float(2.5,3.5)
                        pg.press('enter')
                    return {'status' : False}
                print('이미지 업로드 작업 돈다!')
                while True:
                    print('이미지 업로드 클릭!!')
                    try:
                        wait_float(1.2,1.9)
                        imageBtn = driver.find_element(by=By.CSS_SELECTOR, value=".is-scroll .toolbarBtn.image")
                        imageBtn.click()
                        break
                    except:
                        pass

                    try:
                        wait_float(1.2,1.9)
                        toolbarList = driver.find_element(by=By.CSS_SELECTOR, value=".toolbarList")
                        toolbarList.click()
                    except:
                        pass
                
                wait_float(2.1,2.9)
                
                imagePath = nowPath + f"\etc\img\{workBlogNum}"
                print(imagePath)
                wait_float(0.5, 0.9)
                pyperclip.copy(imagePath)
                wait_float(0.5, 0.9)
                pg.hotkey('ctrl','v')
                wait_float(0.5, 0.9)
                pg.press('enter')
                
                wait_float(0.5, 0.9)
                print(chkAction[1])
                pyperclip.copy(chkAction[1])
                wait_float(0.5, 0.9)
                pg.hotkey('ctrl','v')
                wait_float(0.5, 0.9)
                pg.press('enter')
                wait_float(3.5,4.5)

                # 업로드 된 이미지 갯수 확인 후
                try:
                    imageCountChk = driver.find_elements(by=By.CSS_SELECTOR, value='.se_mediaImage')
                    print(len(imageCountChk))
                    if len(imageCountChk) == imageUploadCount:
                        break
                    else:
                        pg.press('esc')
                except:
                    pass

            imageAlignStatus = False
            while True:
                try:
                    wait_float(0.6, 1.2)
                    toolbarList = driver.find_element(by=By.CSS_SELECTOR, value=".toolbarList")
                    toolbarList.click()
                except:
                    pass

                for align in alignArr:
                    try:
                        wait_float(0.6, 1.2)
                        alignBtn = driver.find_element(by=By.CSS_SELECTOR, value=f".toolbarBtn.{align}")
                        alignBtn.click()
                        alignCenter = driver.find_elements(by=By.CSS_SELECTOR, value=".toolbarBtn.align-center")
                        if len(alignCenter) > 0:
                            imageAlignStatus = True
                            break
                    except:
                        pass
                if imageAlignStatus == True:
                    break
                
            while True:
                pg.moveTo(300,500)
                try:
                    wait_float(0.6, 1.2)
                    nextLineArr = driver.find_elements(by=By.CSS_SELECTOR, value=".__cursorPlaceholder.btn_cursorPlaceholder")
                    nextLineArr[-1].click()
                    pg.scroll(-100)
                    break
                except:
                    pg.scroll(-100)
                    pass
            
            alignStatus = False
            while True:
                try:
                    wait_float(0.6, 1.2)
                    toolbarList = driver.find_element(by=By.CSS_SELECTOR, value=".toolbarList")
                    toolbarList.click()
                except:
                    pass

                for align in alignArr:
                    try:
                        wait_float(0.6, 1.2)
                        alignBtn = driver.find_element(by=By.CSS_SELECTOR, value=f".toolbarBtn.{align}")
                        alignBtn.click()
                        alignCenter = driver.find_elements(by=By.CSS_SELECTOR, value=".toolbarBtn.align-center")
                        if len(alignCenter) > 0:
                            alignStatus = True
                            break
                    except:
                        pass
                if alignStatus == True:
                    break

            continue
        
        if chkAction[0] == 'link':

            linkAction = ""

            if len(chkAction) == 2:
                linkAction = "single"
            elif len(chkAction) > 2:
                linkAction = "multi"

            focus_window('Chrome')

            if linkAction == 'multi':
                keyboard.write(text=chkAction[1], delay=0.05)
                wait_float(0.3,0.9)
                pg.keyDown('shiftleft')
                pg.keyDown('shiftright')
                for linkIndex in range(len(chkAction[1])):
                    pg.hotkey('left')
                    wait_float(0.1,0.3)
                pg.keyUp('shiftleft')
                pg.keyUp('shiftright')

            while True:
                try:
                    wait_float(0.5,1.2)
                    imageBtn = driver.find_element(by=By.CSS_SELECTOR, value=".is-scroll .toolbarBtn.link")
                    imageBtn.click()
                except:
                    pass

                try:
                    wait_float(0.5,1.2)
                    toolbarList = driver.find_element(by=By.CSS_SELECTOR, value=".toolbarList")
                    toolbarList.click()
                except:
                    pass

                try:
                    wait_float(0.5,1.2)
                    inputUrl = driver.find_element(by=By.CSS_SELECTOR, value=".input_url")
                    inputUrl.click()
                    if linkAction == 'multi':
                        keyboard.write(text=chkAction[2], delay=0.05)
                    else:
                        keyboard.write(text=chkAction[1], delay=0.05)
                except:
                    pass

                try:
                    wait_float(0.5,1.2)
                    inputUrlConfirm = driver.find_element(by=By.CSS_SELECTOR, value=".pop_btn.btn_apply")
                    inputUrlConfirm.click()
                    break
                except:
                    pass

            if linkAction == 'multi':
                wait_float(0.3,0.9)
                pg.press("right")
            wait_float(0.3,0.9)
            pg.press("enter")
            wait_float(0.5,0.9)
            pg.press("enter")
            wait_float(0.5,0.9)

            continue

        if lineCount == 0:
            pg.moveTo(300,500)
            pg.scroll(150)
            wait_float(1.2,1.9)
            pg.scroll(150)
            wait_float(1.2,1.9)
            pg.scroll(150)
            wait_float(1.2,1.9)
            # 여기가 제목~~
            while True:
                try:
                    wait_float(0.5,0.9)
                    subjectArea = driver.find_element(by=By.CSS_SELECTOR, value=".se_editView.se_title")
                    subjectArea.click()
                    break
                except:
                    pass

            wait_float(0.5,0.9)
            keyboard.write(text=getline, delay=0.12)
            wait_float(1.2,2.8)
        elif lineCount == 1:
            # 본문 들어가기 시작할때~~~
            focus_window('Chrome')

            while True:
                try:
                    wait_float(1.2,1.9)
                    editorArea = driver.find_element(by=By.CSS_SELECTOR, value=".se_editable")
                    editorArea.click()
                    break
                except:
                    pass

            wait_float(1.2,1.9)

                
            keyboard.write(text=getline, delay=0.12)
            focus_window('Chrome')
            wait_float(0.5,0.9)
            pg.press('enter')
            wait_float(0.5,0.9)
            wait_float(2.5,2.9)
            
            
        elif getline == 'enter':
            focus_window('Chrome')
            wait_float(0.5,0.9)
            pg.press('enter')
            wait_float(0.5,0.9)
        else:
            focus_window('Chrome')
            keyboard.write(text=getline, delay=0.12)
            wait_float(0.3,0.9)
            pg.press('enter')
            wait_float(1.5,1.9)

    if httpsEnter > 0:
        pg.moveTo(300,500)
        wait_float(0.3,0.9)
        pg.scroll(150)
        wait_float(0.3,0.9)
        pg.scroll(150)
        wait_float(0.3,0.9)
        pg.scroll(150)
        wait_float(0.3,0.9)
        pg.scroll(150)
        wait_float(0.3,0.9)
        pg.scroll(150)
        while True:
            print('돈다잉!!!')
            try:
                parents = driver.find_elements(by=By.CSS_SELECTOR, value=".se_editable")
                childs = parents[-1].find_elements(By.XPATH, "./*")
            except:
                continue
            prevTag = ''
            nowTag = ''
            startVal = False
            for idx, child in enumerate(childs):
                print(child.tag_name)
                nowTag = child.tag_name
                if nowTag == 'br' and prevTag == 'br' and startVal == True:
                    wait_float(1.2,1.9)
                    pg.press('down')
                    wait_float(1.2,1.9)
                    pg.press('delete')
                try:
                    if child.tag_name == 'a':
                        startVal = True
                        child.click()
                except:
                    pass
                prevTag = nowTag
                
            break

    subjectAlignStatus = False
    while True:
        pg.moveTo(200,500)
        wait_float(0.3,0.6)
        pg.scroll(300)
        wait_float(0.3,0.6)
        pg.scroll(300)
        wait_float(0.3,0.6)
        pg.scroll(300)
        try:
            wait_float(1.2,1.9)
            subjectArea = driver.find_element(by=By.CSS_SELECTOR, value=".se_editView.se_title")
            subjectArea.click()
        except:
            pass

        try:
            wait_float(1.2,1.9)
            subjectAlignChk = driver.find_elements(by=By.CSS_SELECTOR, value=".toolbarBtn.documentTitleAlign-center")
            print(len(subjectAlignChk))
            if len(subjectAlignChk) > 0:
                subjectAlignChk[0].click()
        except:
            pass


        try:
            wait_float(1.2,1.9)
            subjectAlignChkLeft = driver.find_elements(by=By.CSS_SELECTOR, value=".toolbarBtn.documentTitleAlign-left")
            if len(subjectAlignChkLeft) > 0:
                subjectAlignStatus = True
        except:
            pass

        if subjectAlignStatus == True:
            break


    
         





    # 글 작성 완료 하는 부분!!!
    while True:
        pg.moveTo(200,500)
        pg.scroll(-300)
        for i in range(2):
            pg.scroll(120)
            try:
                wait_float(1.5,2.3)
                applyPost = driver.find_element(by=By.CSS_SELECTOR, value=".btn_applyPost")
                applyPost.click()
            except:
                pass

        try:
            wait_float(1.2,1.9)
            postCompleteChk = driver.find_elements(by=By.CSS_SELECTOR, value=".btn_function._btn_tools")
            if len(postCompleteChk) > 0:
                break
        except:
            pass

    current_url = driver.current_url

    return {'status' : True, 'url' : current_url}

    

            
def naverLogin_mobile(driver, idInfo):

    while True:
        print("로그인 돌아돌아11111!")
        errCount = 0
        while True:
            print("로그인 돌아돌아22222!")
            errCount += 1
            # btn_retry
            print(errCount)
            if errCount > 5:
                focus_window('로그인')
                wait_float(0.5,1.2)
                pg.press('F5')
                errCount = 0

            try:
                print('재시도가 나오는 경우가 있음!!!')
                wait_float(1.2,1.9)
                retryBtn = driver.find_element(by=By.CSS_SELECTOR, value=".btn_retry")
                retryBtn.click()
            except:
                pass

            try:
                print('혹시 모르니 메인 가는거 클릭 한번!!!')
                wait_float(1.2,1.9)
                main = driver.find_element(by=By.CSS_SELECTOR, value=".sch_logo_naver.MM_SEARCH_GO_HOME")
                main.click()
            except:
                pass

            
            
            try:
                print('메인 > 로그인 창까지 이동!')
                wait_float(1.2,1.9)
                aside = driver.find_element(by=By.CSS_SELECTOR, value=".sch_ico_aside")
                aside.click()
            except:
                pass
            
            try:
                wait_float(1.2,1.9)
                aside = driver.find_element(by=By.CSS_SELECTOR, value=".sha_ico_aside")
                aside.click()
            except:
                pass
            
            # 이미 로그인이 되어 있는지 체크~
            
            try:
                print('로그인 버튼 클릭하기!!')
                wait_float(1.2,1.9)
                gotoLogin = driver.find_element(by=By.CSS_SELECTOR, value=".link_user")
                if '로그인' in gotoLogin.text:
                    gotoLogin.click()
                elif '님' in gotoLogin.text:
                    driver.get('https://naver.com')
                    return True
            except:
                pass
            
            try:
                print('로그인 페이지 왔으면 break')
                wait_float(1.2,1.9)
                driver.find_element(by=By.CSS_SELECTOR, value="#id")
                pg.click(50,300)
                break
            except:
                pass

            try:
                wait_float(1.2,1.9)
                closePopup = driver.find_element(by=By.CSS_SELECTOR, value=".lst_btn_close")
                print('새로운 팝업창 클릭 확인!')
                closePopup.click()
            except:
                pass

            

        # 로그인 부분
        
        focus_window('로그인')

        while True:
            print('로그인 하는거니??')
            try:
                wait_float(2.2,2.9)
                pyperclip.copy(idInfo['n_id'])
                id_input = driver.find_element(by=By.CSS_SELECTOR, value="#id")
                id_input.click()
                wait_float(0.4, 0.7)
                pg.hotkey('ctrl', 'a')
                wait_float(0.4, 0.7)
                pg.hotkey('ctrl', 'v')
                wait_float(0.4, 0.7)

                pyperclip.copy(idInfo['n_pwd'])
                pw_input = driver.find_element(by=By.CSS_SELECTOR, value="#pw")
                pw_input.click()
                wait_float(0.4, 0.7)
                pg.hotkey('ctrl', 'a')
                wait_float(0.4, 0.7)
                pg.hotkey('ctrl', 'v')
                wait_float(0.4, 0.7)
                id_input_value = id_input.get_attribute('value')
                if id_input_value:
                    pg.hotkey('enter')
                    wait_float(0.5, 1.0)
                    print('일단 로그인 시도 완료!!')
                    break
                else:
                    continue
            except:
                continue


        loginInfo = {}
        loginInfo['loginErrMessage'] = ""

        while True:

            # 캡챠 걸리면 break 치고 다시~~~~!!
            try:
                wait_float(0.3,0.5)
                captchaWrap = driver.find_elements(by=By.CSS_SELECTOR, value=".captcha_wrap")
                if len(captchaWrap) > 0:
                    print('캡챠 뜸!!')
                    wait_float(3.5,5.5)
                    driver.get('https://m.naver.com')
                    break
            except:
                pass

            # 새로운 환경에서 로그인
            try:
                wait_float(0.3,0.5)
                newDevice = driver.find_elements(by=By.CSS_SELECTOR, value=".btn_white")
                print(f'newDevice : {newDevice}')
                if len(newDevice) > 0:
                    print('새로운 환경에서 로그인')
                    newDevice[0].click()
                    return True
            except:
                pass

            # aside_header가 나오면 로그인 성공임!!!
            try:
                wait_float(0.3,0.5)
                loginSuccess = driver.find_elements(by=By.CSS_SELECTOR, value=".aside_header")
                print(f'loginSuccess : {loginSuccess}')
                if len(loginSuccess) > 0:
                    print('로그인 성공!')
                    driver.get('https://naver.com')
                    return True
            except:
                pass

            try:
                wait_float(0.3,0.5)
                greenBtn = driver.find_element(by=By.CSS_SELECTOR, value=".btn_next")
                print(f'greenBtn : {greenBtn}')
                if greenBtn and "보호조치" in greenBtn.text:
                    print('보호조치 ㅠ')
                    loginInfo['loginErrMessage'] = "보호조치"
                    return False
                elif greenBtn and "대량생성" in greenBtn.text:
                    loginInfo['loginErrMessage'] = "대량생성"
                    return False
                
            except:
                pass

            # try:
            #     wait_float(0.3,0.5)
            #     titleWrap = driver.find_element(by=By.CSS_SELECTOR, value=".title_wrap")
            #     if titleWrap and "휴면" in titleWrap.text:
            #         loginInfo['loginErrMessage'] = "휴면아이디"
            #         return False
            # except:
            #     pass

            try:
                
                wait_float(0.3,0.5)
                errMessage = driver.find_element(by=By.CSS_SELECTOR, value="#error_message")
                print(f'errMessage : {errMessage.text}')
                if errMessage and "자동입력" in errMessage.text:
                    print('캡챠 뜸!!')
                    driver.get('https://m.naver.com')
                    break
            except:
                pass

            try:

                wait_float(0.3,0.5)
                errMessage = driver.find_element(by=By.CSS_SELECTOR, value="#error_message")
                if errMessage and "비밀번호를 잘못" in errMessage.text:
                    loginInfo['loginErrMessage'] = "비번틀림"
                    print('비번 틀림!!')
                    return False
            except:
                pass

            

            try:
                wait_float(0.3,0.5)
                errMessage = driver.find_element(by=By.CSS_SELECTOR, value=".top_title")
                if errMessage and "비정상" in errMessage.text:
                    loginInfo['loginErrMessage'] = "비정상 활동"
                    print('비정상적 활동!!')
                    return False
            except:
                pass



def naver_login(driver, blogInfo):
    while True:
        print('로그인 버튼 클릭하기~~~')
        try:
            wait_float(1.2,1.9)
            loginBtn = driver.find_element(by=By.CSS_SELECTOR, value=".MyView-module__link_login___HpHMW")
            loginBtn.click()
        except:
            pass
        
        try:
            wait_float(0.5,1.2)
            idInputChk = driver.find_element(by=By.CSS_SELECTOR, value="#input_item_id")
            if idInputChk:
                break
        except:
            pass
    focus_window('로그인')
    wait_float(0.3,0.9)
    
    while True:
        print('돌고돌고~~')
        focus_window('Chrome')
        
        try:
            print('로그인 부분')
            # pg.click(250,500)
            inputId = driver.find_element(by=By.CSS_SELECTOR, value="#id")
            fRanVal = random.randrange(20,140)
            sRanVal = random.randrange(80,90)

            print(inputId.location['y'])
            print(inputId.location['x'])
            
            durationRanVal = random.uniform(0.3,0.8)
            # durationRanVal = random.randrange(0.5,1.8)
            pg.moveTo(inputId.location['x'] + fRanVal, inputId.location['y'] + sRanVal, durationRanVal)
            wait_float(0.5,1.2)
            pg.click(inputId.location['x'] + fRanVal, inputId.location['y'] + sRanVal)
            wait_float(0.3,0.9)
            print('아이디 란 마우스 이동 완료')
            inputId.click()
            print('아이디 란 클릭 완료')
            wait_float(0.3,0.9)
            cb.copy(blogInfo['n_id'])
            wait_float(0.3,0.9)
            pg.hotkey('ctrl', 'a')
            wait_float(0.3,0.9)
            pg.hotkey('ctrl', 'v')
            
            print('아이디 붙여넣기 완료')

            inputId = driver.find_element(by=By.CSS_SELECTOR, value="#id")
            
            inputPw = driver.find_element(by=By.CSS_SELECTOR, value="#pw")
            
            fRanVal = random.randrange(20,140)
            sRanVal = random.randrange(80,90)
            
            durationRanVal = random.uniform(0.3,0.8)
            # durationRanVal = random.randrange(0.5,1.8)
            pg.moveTo(inputPw.location['x'] + fRanVal, inputPw.location['y'] + sRanVal, durationRanVal)
            wait_float(0.5,1.2)
            pg.click(inputPw.location['x'] + fRanVal, inputPw.location['y'] + sRanVal)
            wait_float(0.3,0.9)
            print('비번 란 마우스 이동 완료')
            
            inputPw.click()
            print('비번 란 클릭 완료')
            wait_float(0.3,0.9)
            cb.copy(blogInfo['n_pwd'])
            wait_float(0.3,0.9)
            pg.hotkey('ctrl', 'a')
            wait_float(0.3,0.9)
            pg.hotkey('ctrl', 'v')
            print('비번 붙여넣기 완료')
            inputPw = driver.find_element(by=By.CSS_SELECTOR, value="#pw")
            if inputPw.get_attribute('value') != "" and inputId.get_attribute('value') != "":
                print('아이디 비번 둘다 안비있음 체크 완료')
                wait_float(1.2,1.9)
                btnLogin = searchElement('.btn_login',driver)
                btnLogin[0].click()
                wait_float(0.3,0.5)
                pg.press('enter')
                wait_float(3.5,4.5)
        except:
            pass

        try:
            loginBox = driver.find_elements(by=By.CSS_SELECTOR, value="#login_box")
            topSearchWrap = driver.find_elements(by=By.CSS_SELECTOR, value="#topSearchWrap")
            if len(loginBox) == 0 and len(topSearchWrap) > 0:
                print(f'로그인 박스! {len(loginBox)} // 메인 체크! {len(topSearchWrap)}')
                print('로그인 성공했으면 나가기!!')
                break
        except:
            pass

        try:
            loginErrEle1 = driver.find_elements(by=By.CSS_SELECTOR, value=".message_wrap")
            if len(loginErrEle1) > 0:
                print('그 이상한 영수증!!!')
                for i in range(3):
                    fr = 1600    # range : 37 ~ 32767
                    du = 500     # 1000 ms ==1second
                    sd.Beep(fr, du)
                wait_float(4.3,5.5)
                pg.press('F5')
                try:
                    WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#query")))
                except:
                    print('메인으로 갑시다!!')
                    driver.get('https://www.naver.com')
                # complexBool = pg.confirm('에러남!! 직접 풀고 계속 진행??')
                # if complexBool == 'OK':
                #     complexAuthStatus = False
                # else:
                #     complexAuthStatus = True
                #     driver.quit()
                # break
        except:
            pass

        

        wait_float(1.2,1.9)
        try:
            loginErrEle2 = driver.find_elements(by=By.CSS_SELECTOR, value="#divWarning")
            if len(loginErrEle2) > 0:
                print('보호조치 걸림!! return!!!')
                for i in range(3):
                    fr = 1600    # range : 37 ~ 32767
                    du = 500     # 1000 ms ==1second
                    sd.Beep(fr, du)
                driver.quit()
                return False
        except:
            pass

def convert_to_jpeg(image_path, output_path):
    try:
        # 이미지 열기
        img = Image.open(image_path)

        # 파일 형식 확인
        image_format = img.format

        # PNG 또는 JPEG 형식이 아닌 경우 JPEG 형식으로 변환
        if image_format not in ['JPEG', 'PNG']:
            img = img.convert("RGB")
            img.save(output_path, format='JPEG')
            print(f"{image_path}를 JPEG 파일로 변환하여 {output_path}에 저장되었습니다.")
        else:
            change_extension(image_path, 'jpg')
            print(f"{image_path}는 이미 JPEG 또는 PNG 파일입니다. 변환이 필요하지 않습니다.")
    except Exception as e:
        print(f"오류가 발생했습니다: {e}")

def change_extension(file_path, new_extension):
    # 파일의 경로와 확장자를 분리
    root, _ = os.path.splitext(file_path)
    # 새로운 확장자를 추가하여 새 파일 경로 생성
    new_file_path = root + '.' + new_extension
    try:
        # 파일 이름 변경
        os.rename(file_path, new_file_path)
        print(f"{file_path}의 확장자가 {new_extension}으로 변경되었습니다.")
    except OSError as e:
        print(f"파일 이름 변경 중 오류가 발생했습니다: {e}")

# 블로그 글쓰기
def writeBlog(driver,workBlogNum,contentArr):


    getTime = None
    driver.switch_to.default_content()

    driver.get('https://www.naver.com?mobile')

    
    errCount = 0
    while True:
        wait_float(0.3,0.8)
        errCount += 1
        if errCount > 5:
            driver.get('https://www.naver.com?mobile')
            errCount = 0
        try:
            navList = driver.find_elements(by=By.CSS_SELECTOR, value='.service_name')
            if navList:
                break
        except:
            pass
    for mitem in navList:
        if mitem.text == '블로그':
            mitem.click()
            break
    
    driver.close()
    driver.switch_to.window(driver.window_handles[0])

    # 블로그 찾아 들어가기
    errCount = 0
    while True:
        wait_float(2.5,3.3)
        errCount += 1
        if errCount > 3:
            pg.press('F5')
        if errCount > 6:
            errCount = 0
            for i in range(1):
                fr = 1550    # range : 37 ~ 32767
                du = 800     # 1000 ms ==1second
                sd.Beep(fr, du)
            pg.alert('아이디 생성 체크하기!!')
        try:
            menu_my_blog = driver.find_elements(by=By.CSS_SELECTOR, value='.menu_my_blog .item')
            menu_my_blog[1].click()
        except:
            pass
        
        if len(driver.window_handles) == 2:
            break
    
    errCount = 0
    while True:
        wait_float(3.5,5.5)
        errCount += 1
        if errCount > 5:
            errCount = 0
            focus_window('Chrome')
            pg.press('F5')
        print('글쓰기로 들어가기!!!')
        
        try:
            driver.switch_to.window(driver.window_handles[1])
            wait_float(1.2,1.9)
            driver.switch_to.frame('mainFrame')
            print('첫번째 체크!!!')
            focus_window('Chrome')
            writeArea = driver.find_elements(by=By.CSS_SELECTOR, value='.se-component-content')
            writeArea[0].click()
            wait_float(2.5,3.5)
        # except TimeoutException as te:
        #     driver.refresh()  # 예: 새로고침 시도
        #     continue
        except Exception as e:
            print(str(e))
            print('메인 프레임 에러!!!')
            pass

        try:
            wait_float(0.5,1.2)
            focus_window('Chrome')
            writeArea = driver.find_elements(by=By.CSS_SELECTOR, value='.se-component-content')
            writeArea[0].click()
            wait_float(2.5,3.5)
        except Exception as e:
            print(str(e))
            print('제목 클릭 에러!!!')
            pass

        try:
            helpCloseBtn = driver.find_element(by=By.CSS_SELECTOR, value=".se-help-panel-close-button")
            helpCloseBtn.click()
        except Exception as e:
            pass

       
        try:
            preWritePopup = driver.find_element(by=By.CSS_SELECTOR, value=".se-popup-button-cancel")
            preWritePopup.click()
            wait_float(0.5,0.9)
        except Exception as e:
            pass
        
        try:
            subject_area = writeArea[0].find_element(by=By.CSS_SELECTOR, value=".se-section-documentTitle")
            subject_area.click()
            break
        except Exception as e:
            pass

    lineCount = -1
    while True:

        lineCount += 1

        if lineCount > len(contentArr):
            break

        focus_window('Chrome')

        # 리스트 범위 내에거 찾아야함!!
        try:
            getline = contentArr[lineCount]
        except:
            continue


        chkAction = getline.split('|')
        if chkAction[0] == 'img_line':
            nowPath = os.getcwd()
            
            try:
                if chkAction[1] == 'enter':
                    continue
                elif chkAction[1]:
                    pass
                else:
                    continue
            except:
                continue

            while True:
                print('이미지 추가 버튼 클릭 부분!!')
                wait_float(0.5,1.2)
                try:
                    img_btn = driver.find_element(by=By.CSS_SELECTOR, value='.se-image-toolbar-button')
                    img_btn.click()
                    break
                except:
                    pass
            wait_float(1.5,2.3)
        
            imagePath = nowPath + f"\etc\img\{workBlogNum}"
            wait_float(1.5, 2.2)
            pyperclip.copy(imagePath)
            wait_float(0.5, 0.9)
            pg.hotkey('ctrl','v')
            wait_float(0.5, 0.9)
            pg.press('enter')
            
            wait_float(0.9, 1.6)
            pyperclip.copy(chkAction[1])
            wait_float(0.5, 0.9)
            pg.hotkey('ctrl','v')
            wait_float(0.5, 0.9)
            pg.press('enter')
            wait_float(3.5,4.5)

            # 이미지 업로드 에러? 났을때!!!
            try:
                imgErrorChkBtn = driver.find_element(by=By.CSS_SELECTOR, value='.se-popup-button.se-popup-button-confirm')
                imgErrorChkBtn.click()
                wait_float(1.5,2.5)
            except:
                pass

            continue
        
        if chkAction[0] == 'link':

            if len(chkAction) == 3:
                focus_window('Chrome')
                keyboard.write(text=chkAction[1], delay=0.05)
                wait_float(0.3,0.9)
                pg.keyDown('shiftleft')
                pg.keyDown('shiftright')
                for linkIndex in range(len(chkAction[1])):
                    pg.hotkey('left')
                    wait_float(0.1,0.3)
                pg.keyUp('shiftleft')
                pg.keyUp('shiftright')

                while True:
                    try:
                        linkBtn = driver.find_element(by=By.CSS_SELECTOR, value='.se-link-toolbar-button')
                        linkBtn.click()
                    except:
                        pass

                    try:
                        linkBox = driver.find_element(by=By.CSS_SELECTOR, value='.se-custom-layer-option-link')
                        if linkBox:
                            break
                    except:
                        pass
                
                while True:
                    try:
                        linkInput = driver.find_element(by=By.CSS_SELECTOR, value='.se-custom-layer-link-input')
                        linkInput.click()

                        keyboard.write(text=chkAction[2], delay=0.05)

                        linkChkBtn = driver.find_element(by=By.CSS_SELECTOR, value='.se-custom-layer-link-apply-button')
                        linkChkBtn.click()
                        
                    except:
                        pass

                    try:
                        linkBoxs = driver.find_elements(by=By.CSS_SELECTOR, value='.se-custom-layer-option-link')
                        if len(linkBoxs) == 0:
                            break
                    except:
                        pass
                wait_float(0.3,0.9)
                pg.press("right")
                wait_float(0.3,0.9)
                pg.press("enter")
                wait_float(0.5,0.9)
                pg.press("enter")
                wait_float(0.5,0.9)
            elif len(chkAction) == 2:
                keyboard.write(text=chkAction[1], delay=0.1)
                wait_float(1.2,1.9)
                pg.press('enter')
                wait_float(1.2,1.9)
                pg.press('enter')
                while True:
                    print('여기가 막히나?!?!?!')
                    wait_float(1.2,1.9)
                    try:
                        linkBoxList = driver.find_elements(by=By.CSS_SELECTOR, value='.se-oglink-frame')
                        print(len(linkBoxList))
                        if len(linkBoxList) == 0:
                            continue
                        else:
                            break
                    except Exception as e:
                        print(str(e))
                        pass
                while True:
                    print('저긴 넘어갔고 여기가 맞는데;;;;')
                    wait_float(1.2,1.9)
                    try:
                        linkBoxList = driver.find_elements(by=By.CSS_SELECTOR, value='.se-oglink-frame')
                        if len(linkBoxList) == 0:
                            break
                    except Exception as e:
                        pass
                    try:
                        linkBox = driver.find_element(by=By.CSS_SELECTOR, value='.se-oglink-thumbnail-frame')
                        linkBox.click()
                    except:
                        pass
                    try:
                        boxDeleteBtn = driver.find_element(by=By.CSS_SELECTOR, value='.se-delete-toolbar-button')
                        boxDeleteBtn.click()
                    except:
                        pass
                
                wait_float(1.2,1.9)
                pg.press('enter')
                wait_float(1.2,1.9)
                pg.press('enter')

            continue

        if lineCount == 0:
            # 여기가 제목~~
            writeArea[0].click()
            keyboard.write(text=getline, delay=0.05)
            wait_float(1.2,2.8)
        elif lineCount == 1:
            # 본문 들어가기 시작할때~~~
            focus_window('Chrome')
            writeArea[1].click()
            while True:
                try:
                    alignBoxChk = driver.find_elements(by=By.CSS_SELECTOR, value='.se-align-center-toolbar-button')
                    if alignBoxChk:
                        break
                except:
                    pass
                
                try:
                    alignBoxChk = driver.find_elements(by=By.CSS_SELECTOR, value='.se-align-left-toolbar-button')
                    alignBoxChk[0].click()
                    alignBtnList = driver.find_elements(by=By.CSS_SELECTOR, value=".se-toolbar-option.se-toolbar-option-align button")
                    wait_float(0.5,1.2)
                    alignBtnList[1].click()
                    wait_float(0.5,1.2)
                    break
                except:
                    pass
                
            keyboard.write(text=getline, delay=0.05)
            focus_window('Chrome')
            wait_float(0.5,0.9)
            pg.press('enter')
            wait_float(0.5,0.9)
            wait_float(2.5,2.9)
            
            
        elif getline == 'enter':
            focus_window('Chrome')
            wait_float(0.5,0.9)
            pg.press('enter')
            wait_float(0.5,0.9)
        else:
            focus_window('Chrome')
            keyboard.write(text=getline, delay=0.05)
            wait_float(0.3,0.9)
            pg.press('enter')
            wait_float(1.5,1.9)
            if 'http' in getline:
                wait_float(3.5,3.9)

            try:
                linkBox = driver.find_element(by=By.CSS_SELECTOR, value='.se-module-oglink')
                if linkBox:
                    linkBox.click()
                    wait_float(0.5,0.9)
                    delBtns = driver.find_elements(by=By.CSS_SELECTOR, value='.se-delete-toolbar-button')
                    delBtns[-1].click()
            except Exception as e:
                pass
    

    print('글 작성 끝남!!!!!!!!!!!')
    try:

        # 발행버튼 누르기
        while True:
            print('발행버튼 클릭~~~')
            wait_float(0.5,1.2)
            try:
                publichBtn = driver.find_element(By.CSS_SELECTOR, '[class*="publish_btn__"]')
                publichBtn.click()
                break
            except:
                pass

        # 공감 체크 되어 있는지 확인 및 안되어 있으면 체크
        while True:
            print('공감체크~~~')
            wait_float(0.5,1.2)
            try:
                gongamChkbox = driver.find_element(by=By.CSS_SELECTOR, value='#publish-option-sympathy')
                gongamChkboxLabel = gongamChkbox.find_element(By.XPATH, '..')
                if gongamChkbox.is_selected():
                    break
                else:
                    gongamChkboxLabel.click()
            except Exception as e:
                print(str(e))
                pass

        wait_float(1.5,2.5)

        while True:
            print('글 발행 완료 GOGO~~~')
            wait_float(0.3,0.5)
            try:
                confirmBtn = driver.find_element(By.CSS_SELECTOR, '[class*="confirm_btn__"]')
                confirmBtn.click()
                print('글 발행 클릭 완료!!')
                break
            except TimeoutException as te:
                driver.refresh()  # 예: 새로고침 시도
                continue
            except Exception as e:
                pass

        wait_float(3.5,5.5)

        print('글 발행 완료!!!! SUCCESS!!!!!')

    except Exception as e:
        for i in range(4):
            fr = 1800    # range : 37 ~ 32767
            du = 400     # 1000 ms ==1second
            sd.Beep(fr, du)
        pg.alert(e)
        pass
        
    
    wait_float(3.5,5.5)
    

    try:
        miniPopup = driver.find_element(by=By.CSS_SELECTOR, value="#floatingda_content")
        miniPopupClose = miniPopup.find_element(by=By.CSS_SELECTOR, value="button")
        driver.execute_script("arguments[0].scrollIntoView();", miniPopupClose)
        wait_float(0.5,1.2)
        miniPopupClose.click()
        wait_float(0.5,1.2)
    except:
        pass

    writePostLink = ""
    while True:

        try:
            closeLayer = driver.find_element(by=By.CSS_SELECTOR, value=".btn_close._closeLayer")
            closeLayer.click()
        except:
            pass

        try:
            getUrl = searchElement('._transPosition',driver)
            getUrl[0].click()
            wait_float(1.5,2.5)
            pg.press('enter')
            wait_float(1.5,2.5)
            writePostLink = pyperclip.paste()
            if writePostLink != "":
                break
            
        except Exception as e:
            print(str(e))
            pass
        
    driver.switch_to.default_content()
    while True:
        if len(driver.window_handles) > 1:
            driver.switch_to.window(driver.window_handles[1])
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
        else:
            break
        wait_float(0.5,0.9)
    
    wait_float(2.2,2.9)
    
    try:
        closeModal = driver.find_element(by=By.CSS_SELECTOR, value=".label_not_see._btn_no_more_show")
        if closeModal:
            closeModal.click()
    except:
        pass
    
    try:
        goToNaverMain = searchElement('.link_naver',driver)
        goToNaverMain[0].click()
        driver.get('https://www.naver.com')
    except:
        driver.get('https://www.naver.com')

    return {'status' : True, 'url' : writePostLink}
    
    



# 공감 순회하기
def allowListVisit(driver):
    
    while True:
        try:
            navList = driver.find_elements(by=By.CSS_SELECTOR, value='.service_name')
            if navList:
                break
        except:
            pass
    for mitem in navList:
        if mitem.text == '블로그':
            mitem.click()
            break
    
    driver.close()
    driver.switch_to.window(driver.window_handles[0])
    
    
    menu_my_blog = searchElement('.menu_my_blog .item',driver)
    menu_my_blog[0].click()
    
    
    
    workPostCount = 0
    compareDate = ''
    while True:
        
        conVal = ''
        while True:
            print('팝업 없애고 목록 닫기 체크!')
            driver.switch_to.default_content()
            driver.switch_to.window(driver.window_handles[1])
            driver.switch_to.frame('mainFrame')
            
            wait_float(1.5,2.5)
            try:
                closePopupBtn = driver.find_element(by=By.CSS_SELECTOR, value="#not_see")
                closePopupBtn.click()
            except:
                pass
            
            try:
                closePopupBtn = driver.find_element(by=By.CSS_SELECTOR, value=".popup_da_btn_area ._btn_close")
                closePopupBtn.click()
            except:
                pass
            
            try:
                closePopupBtn = driver.find_element(by=By.CSS_SELECTOR, value=".moment_event_da-label._btn_no_more_show")
                closePopupBtn.click()
            except:
                pass
            
            try:
                postListOpenBtn = driver.find_element(by=By.CSS_SELECTOR, value="#toplistSpanBlind")
                print(postListOpenBtn.text)
                if postListOpenBtn.text == '목록닫기':
                    break
                else:
                    postListOpenBtn.click()
            except:
                pass
                

        while True:
            print('작업할 포스팅 클릭 및 링크따기 작업중!')
            try:
                wait_float(0.7,1.2)
                print('여기가 중요한데?!?!?!')
                workPostCount += 1
                
                postListWrap = driver.find_element(by=By.CSS_SELECTOR, value=".wrap_blog2_categorylist")
                # workPost = postListWrap.find_elements(by=By.CSS_SELECTOR, value=".ell2.pcol2")
                workPostTr = postListWrap.find_elements(by=By.CSS_SELECTOR, value=".blog2_categorylist tr")
                
                workPostDate = workPostTr[workPostCount].find_element(by=By.CSS_SELECTOR, value=".date")
                
                if compareDate == '' or compareDate == workPostDate.text:
                    print('정상작업~~~')
                    compareDate = workPostDate.text
                    workPostLink = workPostTr[workPostCount].find_element(by=By.CSS_SELECTOR, value=".ell2.pcol2")
                    workPostLink.click()
                else:
                    print('여기서 끝내기!!!!')
                    driver.close()
                    wait_float(0.3,0.9)
                    driver.switch_to.window(driver.window_handles[0])
                    wait_float(0.3,0.9)
                    driver.switch_to.default_content()

                    while True:
                        try:
                            goToMain = driver.find_element(by=By.CSS_SELECTOR, value='.link_naver')
                            goToMain.click()
                        except:
                            pass
                        
                        try:
                            WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#query")))
                            return
                        except:
                            driver.get('https://www.naver.com')
            except:
                continue
            
            try:
                subjectEle = driver.find_element(by=By.CSS_SELECTOR, value=".se-title-text")
            except:
                continue
            
            
            try:
                getUrl = driver.find_element(by=By.CSS_SELECTOR, value="._transPosition")
                getUrl.click()
                wait_float(0.5,1.2)
                pg.press('enter')
                wait_float(0.5,1.2)
                pg.press('enter')
                nowBlogLink = pyperclip.paste()
                nowBlogLinkSplit = nowBlogLink.split('/')
                print(nowBlogLinkSplit)
                if nowBlogLinkSplit[-1].isdigit():
                    break
            except:
                pass
            
        print('작업할 포스팅 클릭 및 링크따기 작업 완료!')
        # openVisitListBtn = searchElement(f'#Sympathy{nowBlogLinkSplit[-1]} .bu_arr',driver)
        # openVisitListBtn[0].click()
        
        
        
        errCount = 0
        while True:
            errCount += 1
            if errCount > 5:
                wait_float(0.5,1.5)
                driver.close()
                wait_float(0.3,0.9)
                driver.switch_to.window(driver.window_handles[0])
                wait_float(0.3,0.9)
                driver.get('https://www.naver.com')

                return
            print('공감 리스트 클릭해서 공감할 리스트 얻기!')
            try:
                wait_float(1.5,2.5)
                openVisitListBtn = driver.find_element(by=By.CSS_SELECTOR, value=f'#Sympathy{nowBlogLinkSplit[-1]} .bu_arr')
                openVisitListBtn.click()
            except:
                pass
                
            try:
                driver.switch_to.frame(f'sympathyFrm{nowBlogLinkSplit[-1]}')
                wait_float(0.5,1.2)
                visitListWrap = driver.find_element(by=By.CSS_SELECTOR, value='.wrap_blog2_sympathy')
                visitList = visitListWrap.find_elements(by=By.CSS_SELECTOR, value=".area_profile")
                
                print(f"공감 받은 갯수는? {len(visitList)}")
                if visitList:
                    break
            except:
                pass
            
            try:
                noGonggam = driver.find_element(by=By.CSS_SELECTOR, value='.no_sympathy.pcol2')
                if noGonggam:
                    conVal == 'ok'
                    break
            except:
                pass
            
            
        if conVal == 'ok':
            continue
        
        for visitCount in range(len(visitList)):
            print(f'{visitCount}번째 이웃순방 작업 시작!!')
            noWork = ''
            
            while True:
                try:
                    driver.switch_to.default_content()
                    wait_float(0.3,0.9)
                    driver.switch_to.frame('mainFrame')
                    wait_float(0.3,0.9)
                    driver.switch_to.frame(f'sympathyFrm{nowBlogLinkSplit[-1]}')
                    wait_float(0.3,0.9)
                    
                    visitList = driver.find_elements(by=By.CSS_SELECTOR, value='.wrap_blog2_sympathy .nick')
                    visitList[visitCount].click()
                    wait_float(0.7,1.2)
                    driver.switch_to.window(driver.window_handles[2])
                    break
                except:
                    pass
                
                
            # 여기서 블로그 말고 프롤로그면 블로그 클릭하게 하기
            no_posting = ''
            while True:
                print("프롤로그면 블로그 클릭!")
                
                try:
                    wait_float(0.7,1.2)
                    driver.switch_to.default_content()
                    badBlogChk = driver.find_element(by=By.CSS_SELECTOR, value=".desc2")
                    if '제한' in badBlogChk.text:
                        no_posting == 'on'
                        break
                except:
                    pass
                
                
                try:
                    wait_float(0.7,1.2)
                    driver.switch_to.default_content()
                    driver.switch_to.frame('mainFrame')
                    blogMenuChk = driver.find_elements(by=By.CSS_SELECTOR, value="#blog-menu .menu1 li a")
                    if(len(blogMenuChk) > 1):
                        blogMenuChk[1].click()
                except:
                    pass
                
                try:
                    wait_float(0.7,1.2)
                    postListOpenBtn = driver.find_element(by=By.CSS_SELECTOR, value="#toplistSpanBlind")
                    break
                except:
                    pass
                
                try:
                    wait_float(0.7,1.2)
                    postListOpenBtn = driver.find_element(by=By.CSS_SELECTOR, value=".new_blog_inner2")
                    no_posting == 'on'
                    break
                except:
                    pass
                
            if no_posting == 'on':
                wait_float(0.5,1.5)
                driver.close()
                wait_float(0.3,0.9)
                driver.switch_to.window(driver.window_handles[1])
                wait_float(0.3,0.9)
                continue
            
            # 가장 먼저 포스팅 갯수 찾기! 포스팅 갯수 10개 미만이면 패스~~
            
            errCount = 0
            while True:
                errCount += 1
                print(errCount)
                if errCount > 7:
                    noWork = 'on'
                    break
                
                wait_float(0.7,1.2)
                try:
                    print('전체보기 열기')
                    openView = driver.find_element(by=By.CSS_SELECTOR, value="._viewMore")
                    openView.click()
                except:
                    pass
                
                try:
                    print("포스팅 갯수 찾기!")
                    postingCountEle = driver.find_element(by=By.CSS_SELECTOR, value=".num.cm-col1")
                    postingCount = re.sub(r'[^0-9]', '', postingCountEle.text)
                    print(f"작업중 블로그의 포스팅 갯수는?! {int(postingCount)}")
                    
                    if int(postingCount) < 10:
                        noWork = 'on'
                    else:
                        postingCountEle.click()
                except:
                    pass
                
                
                try:
                    pg.moveTo(500,500)
                    pg.scroll(100)
                    print('전체보기 클릭')
                    allView = driver.find_element(by=By.CSS_SELECTOR, value="#category0")
                    allView.click()
                    break
                except:
                    pass
            
            if noWork == 'on':
                wait_float(0.5,1.5)
                driver.close()
                wait_float(0.3,0.9)
                driver.switch_to.window(driver.window_handles[1])
                wait_float(0.3,0.9)
                continue
            
            
            postChkStatus = ''
            while True:
                
                try:
                    print("가장 최근 글 클릭!")
                    wait_float(0.7,1.2)
                    postListWrap = driver.find_element(by=By.CSS_SELECTOR, value=".blog2_categorylist")
                    workPost = postListWrap.find_element(by=By.CSS_SELECTOR, value=".ell2.pcol2")
                    workPost.click()
                    postChkStatus = "latest"
                except:
                    print('글 클릭 에러남?!?!')
                    pass
                
                
                if postChkStatus == 'latest':
                    print('latest는 들어옴??')
                    try:
                        wait_float(1.5,2.2)
                        subjectEle = driver.find_element(by=By.CSS_SELECTOR, value=".se-title-text")
                        if subjectEle:
                            break
                    except:
                        pass
                    
                    try:
                        wait_float(1.2,1.9)
                        subjectEle = driver.find_element(by=By.CSS_SELECTOR, value=".se_textView")
                        if subjectEle:
                            break
                    except:
                        pass
                    
                    try:
                        wait_float(1.2,1.9)
                        subjectEle = driver.find_element(by=By.CSS_SELECTOR, value=".htitle")
                        if subjectEle:
                            break
                    except:
                        pass
                    
                    
                
                try:
                    print("목록 열기 클릭!!")
                    postListOpenBtn = driver.find_element(by=By.CSS_SELECTOR, value="#toplistSpanBlind")
                    if postListOpenBtn.text == '목록닫기':
                        pass
                    else:
                        postListOpenBtn.click()
                except:
                    pass
                
            try:
                if "[공유]" in subjectEle.text:
                    print('스크랩 블로그다~~ 돌아가자~~~')
                    wait_float(0.5,1.5)
                    driver.close()
                    wait_float(0.3,0.9)
                    driver.switch_to.window(driver.window_handles[1])
                    wait_float(0.3,0.9)
                    continue
            except Exception as e:
                pg.alert(e)
                
                
            noSearchCount = 0
            while True:
                noSearchCount += 1
                if noSearchCount > 5:
                    break
                print('공감 돌아돌아')
                try:
                    wait_float(1.2,1.9)
                    gongamSuccess = driver.find_element(by=By.CSS_SELECTOR, value=".u_likeit_list_btn._button.pcol2.on")
                    print('공감 성공!!!')
                    if gongamSuccess:
                        break
                except:
                    pass
                
                try:
                    wait_float(1.2,1.9)
                    gongamBtn = driver.find_element(by=By.CSS_SELECTOR, value=".u_ico._icon.pcol3")
                    gongamBtn.click()
                except:
                    pass
            # while True:
            #     print("공감 버튼 클릭!")
            #     try:
            #         gongamBtnWrap = driver.find_element(by=By.CSS_SELECTOR, value='.post-btn.post_btn2')
                    
            #         gongamBtn = gongamBtnWrap.find_element(by=By.CSS_SELECTOR, value='.u_likeit_list_btn._button.pcol2')
            #         getGonggamStatus = gongamBtn.get_attribute('aria-pressed')
            #         print(getGonggamStatus)
            #         wait_float(0.3,0.9)
            #         if getGonggamStatus == 'false':
            #             gongamBtnSub = gongamBtnWrap.find_element(by=By.CSS_SELECTOR, value='.u_likeit_list_module._reactionModule')
                        
            #             print(gongamBtnSub.text)
            #             driver.execute_script("arguments[0].scrollIntoView();", gongamBtnSub)
                        
            #             wait_float(0.3,0.9)
            #             pg.scroll(600)
            #             wait_float(0.3,0.9)
            #             gongamBtnSub.click()
            #             wait_float(1.5,2.2)
                        
            #             fr = 1800    # range : 37 ~ 32767
            #             du = 400     # 1000 ms ==1second
            #             sd.Beep(fr, du)
                        
            #             pg.alert('대기!!')
            #         break
            #     except:
            #         pass

                
                
            wait_float(0.5,1.5)
            driver.close()
            wait_float(0.3,0.9)
            driver.switch_to.window(driver.window_handles[1])
            wait_float(0.3,0.9)


def addNeighborWork(driver):
    driver.switch_to.default_content()
    neighborEx = load_workbook('./etc/neighbor_list.xlsx')
    neighborSheet = neighborEx.active

    neighborAllCount = 0
    while True:
        neighborAllCount += 1
        if neighborSheet.cell(neighborAllCount, 1).value is None:
            break
    
    ranBlogValArr = random.sample(range(1,neighborAllCount),10)

    for ranVal in ranBlogValArr:
        getLink = neighborSheet.cell(ranVal, 1).value
        driver.get(getLink)

        no_posting = ''
        while True:
            print("프롤로그면 블로그 클릭!")
            try:
                wait_float(0.7,1.2)
                driver.switch_to.default_content()
                driver.switch_to.frame('mainFrame')
                blogMenuChk = driver.find_elements(by=By.CSS_SELECTOR, value="#blog-menu .menu1 li a")
                for blogMenu in blogMenuChk:
                    if blogMenu.text == '블로그':
                        blogMenu.click()
                # if(len(blogMenuChk) > 1):
                #     blogMenuChk[1].click()
            except:
                pass
            
            try:
                wait_float(0.7,1.2)
                postListOpenBtn = driver.find_element(by=By.CSS_SELECTOR, value="#toplistSpanBlind")
                break
            except:
                pass
            
            try:
                wait_float(0.7,1.2)
                postListOpenBtn = driver.find_element(by=By.CSS_SELECTOR, value=".new_blog_inner2")
                no_posting == 'on'
                break
            except:
                pass
        
        # 가장 먼저 포스팅 갯수 찾기! 포스팅 갯수 10개 미만이면 패스~~
        noWork = ''
        while True:
            print("포스팅 갯수 찾기!")
            wait_float(0.7,1.2)
            try:
                
                postingCountEle = driver.find_element(by=By.CSS_SELECTOR, value=".num.cm-col1")
                postingCount = re.sub(r'[^0-9]', '', postingCountEle.text)
                print(f"작업중 블로그의 포스팅 갯수는?! {int(postingCount)}")
                
                if int(postingCount) < 10:
                    noWork = 'on'
                    break
            except:
                pass

            try:
                allview = driver.find_element(by=By.CSS_SELECTOR, value=".cm-head.cm_cur._viewMore")
                allview.click()
                if int(postingCount) > 10:
                    break
            except:
                pass

        
        if noWork == 'on':
            neighborSheet.cell(ranVal, 2).value = '비정상 블로그'
            neighborEx.save('./etc/neighbor_list.xlsx')
            continue

        neighborAlready = False

        while True:

            try:
                wait_float(1.2,1.9)
                addNeighborBtn = driver.find_element(by=By.CSS_SELECTOR, value=".btn_add_nb._addBuddyPop._rosRestrictAll")
                if addNeighborBtn.text != '이웃추가':
                    neighborAlready = True
                    break
                else:
                    addNeighborBtn.click()
            except:
                pass
            
            if len(driver.window_handles) > 1:
                break

        
        if neighborAlready:
            continue

        while True:
            driver.switch_to.window(driver.window_handles[1])
            try:
                wait_float(1.2,1.9)
                btnbtn1 = driver.find_element(by=By.CSS_SELECTOR, value=".button_next._buddyAddNext")
                btnbtn1.click()
            except:
                pass
            
            try:
                wait_float(1.2,1.9)
                btnbtn1 = driver.find_element(by=By.CSS_SELECTOR, value=".button_next._addBuddy")
                btnbtn1.click()
            except:
                pass
            
            try:
                wait_float(1.2,1.9)
                btnbtn1 = driver.find_element(by=By.CSS_SELECTOR, value=".area_button .button_close")
                btnbtn1.click()
            except:
                pass
            
            if len(driver.window_handles) == 1:
                driver.switch_to.window(driver.window_handles[0])
                break




def visitNeighborWork(driver):
    
    while True:
        try:
            navList = driver.find_elements(by=By.CSS_SELECTOR, value='.service_name')
            if navList:
                break
        except:
            pass
    for mitem in navList:
        if mitem.text == '블로그':
            mitem.click()
            break
    
    driver.close()
    driver.switch_to.window(driver.window_handles[0])


    duplicateArr = []

    # 페이징 때문에 while문 돌리기
    pagingNum = -1
    while True:
        if len(duplicateArr) >= 10:
            break
            
        pagingNum += 1



        while True:
            # 아래 for문이 끝났을때 main_frame이 적용 되었을것을 방지

            try:
                driver.switch_to.window(driver.window_handles[0])
                driver.switch_to.default_content()
            except:
                pass
            try:
                wait_float(0.5,1.2)
                pagination = driver.find_elements(by=By.CSS_SELECTOR, value=".wrap_thumbnail_post_list .pagination span a")
            except:
                pass

            try:
                wait_float(0.5,1.2)
                pagination[pagingNum].click()
                break
            except:
                pass

        

        # 한개 페이지 내 postList 찾기
        while True:
            print('한개 페이지 내 postList 찾기')
            try:
                wait_float(0.8,1.5)
                postListWrap = driver.find_element(by=By.CSS_SELECTOR, value=".list_post_article")
            except:
                pass

            try:
                wait_float(0.8,1.5)
                postList = postListWrap.find_elements(by=By.CSS_SELECTOR, value=".info_post")
                if postList:
                    break
            except:
                pass
            
        for post in postList:
            
            try:
                print(f"현재 작업 갯수는? {len(duplicateArr)}")
                if len(duplicateArr) >= 10:
                    break

                chkLike = False
                chkReply = False
                while True:
                    print('메인에서 공감 / 댓글 검증')
                    
                    try:
                        wait_float(0.5,1.2)
                        driver.switch_to.window(driver.window_handles[0])
                        driver.switch_to.default_content()
                    except:
                        continue

                    try:
                        wait_float(0.5,1.2)
                        chkLikeEle = post.find_element(by=By.CSS_SELECTOR, value=".comments .like")
                        if chkLikeEle:
                            chkLike = True
                    except:
                        pass

                    try:
                        wait_float(0.5,1.2)
                        chkReplyEle = post.find_element(by=By.CSS_SELECTOR, value=".comments .reply")
                        if chkReplyEle:
                            chkReply = True
                    except:
                        pass

                    if chkLike or chkReply:
                        break




                while True:
                    print('중복된 블로그인지 체크')
                    try:
                        wait_float(0.5,1.2)
                        getAuthor = post.find_element(by=By.CSS_SELECTOR, value=".name_author")
                        print(getAuthor)
                        break
                    except:
                        pass
                if getAuthor.text in duplicateArr:
                    print('중복된 블로그!! 패스!!')
                    continue
                
                duplicateArr.append(getAuthor.text)

                while True:
                    print('블로그 클릭!!')
                    try:
                        wait_float(0.5,1.2)
                        postTitleLink = post.find_element(by=By.CSS_SELECTOR, value=".title_post")
                        postTitleLink.click()
                    except:
                        pass

                    if len(driver.window_handles) > 1:
                        driver.switch_to.window(driver.window_handles[1])
                        break
                
                print(chkLike)
                gonganStatus = False
                if chkLike:
                    while True:
                        print('공감 클릭 돌아돌아')
                        driver.switch_to.window(driver.window_handles[1])
                        driver.switch_to.frame('mainFrame')
                        
                        try:
                            print('공감 체크 작업 시작!')
                            wait_float(1.2,1.9)
                            gongamBtns = driver.find_elements(by=By.CSS_SELECTOR, value=".u_likeit_list_btn._button.pcol2")
                            for gongamBtn in gongamBtns:
                                if gongamBtn.get_attribute('data-log') == 'lik.llike|lik.lunlike':
                                    print('공감 체크할 요소 확인 완료!')
                                    if gongamBtn.get_attribute('aria-pressed') == 'true':
                                        gonganStatus = True
                                        print(f'공감 체크할 요소 상태는? {gonganStatus}')
                            
                            if gonganStatus == True:
                                break        
                            wait_float(1.2,1.9)
                        except:
                            pass
                        
                        try:
                            print(f'공감 클릭하기 들어옴!! 공감 체크할 요소 상태는? {gonganStatus}')
                            wait_float(1.2,1.9)
                            gongamBtns = driver.find_elements(by=By.CSS_SELECTOR, value=".u_likeit_list_btn._button.pcol2")
                            for gongamBtn in gongamBtns:
                                if gongamBtn.get_attribute('data-log') == 'lik.llike|lik.lunlike' and gonganStatus == False:
                                    print('공감 클릭할 요소 찾음!')
                                    gongamBtn.click()
                                    print('공감 클릭 완료!')
                            wait_float(1.2,1.9)
                        except:
                            pass

                
                print('공감 작업 완료~~~~~~~~~~~~~!!!')
                goChkRan = random.randrange(0,2)
                # goChkRan = 0
                print(f"글쓰기 랜덤값은? : {goChkRan} (0이면 댓글 쓰기)")
                
                
                if chkReply and goChkRan == 0:
                    errChkCount = 0
                    while True:
                        errChkCount += 1
                        if errChkCount == 5:
                            pg.press('F5')
                        elif errChkCount > 9:
                            break
                        print('댓글열기~~~ 돌아돌아')
                        
                        
                        
                        try:
                            wait_float(1.2,1.9)
                            commentToggleBtn = driver.find_element(by=By.CSS_SELECTOR, value=".btn_arr")
                            commentToggleBtn.click()
                            print('댓글열기~~~ 첫번째 시도!!')
                        except:
                            pass
                        
                        try:
                            wait_float(1.2,1.9)
                            commentToggleBtn = driver.find_element(by=By.CSS_SELECTOR, value=".area_comment.pcol3")
                            commentToggleBtn.click()
                            print('댓글열기~~~ 두번째 시도!!')
                        except:
                            pass

                        try:
                            wait_float(1.2,1.9)
                            commentToggleBtn = driver.find_element(by=By.CSS_SELECTOR, value=".area_comment.pcol2")
                            commentToggleBtn.click()
                            print('댓글열기~~~ 세번째 시도!!')
                        except:
                            pass

                        try:
                            wait_float(1.2,1.9)
                            commentTextArea = driver.find_element(by=By.CSS_SELECTOR, value=".u_cbox_guide")
                            commentTextArea.click()
                            print('댓글 area 클릭!!!')
                            break
                        except:
                            pass
                    print('댓글 작성 준비 완료~~!!!')
                    try:
                        with open(f'./etc/blog_reply_list.txt', 'rt', encoding='UTF8') as f:
                            getLines = f.readlines()
                    except:
                        with open(f'./etc/blog_reply_list.txt', 'r') as f:
                            getLines = f.readlines()
                            
                    
                    getReplyRanVal = random.randrange(0,len(getLines))
                    print('댓글 불러오기 완료')
                    
                    keyboard.write(text=getLines[getReplyRanVal], delay=0.05)
                    print('댓글 작성 완료')
                    
                    try:
                        wait_float(1.2,1.9)
                        replySuccessBtn = driver.find_element(by=By.CSS_SELECTOR, value=".u_cbox_btn_upload")
                        replySuccessBtn.click()
                        print('댓글 쓰기 버튼 클릭 완료')
                    except:
                        pass
                
                print('전체 완료! 다음거 준비!!!!!!!!!!')
                wait_float(3.5,4.5)
                driver.close()
                wait_float(3.5,4.5)
            except Exception as e:
                print(e)
                wait_float(1.2,1.9)
                pg.press('enter')
                wait_float(1.2,1.9)
                pg.press('enter')
                wait_float(1.2,1.9)
                if len(driver.window_handles) > 1:
                    driver.switch_to.window(driver.window_handles[1])
                    driver.close()
                pass
            
    

            








    


    pass

# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>함수 시작염


def changeIp():
    try:
        print('아이피 변경 언제??')
        os.system('adb server start')
        client = AdbClient(host="127.0.0.1", port=5037)
        device = client.devices()  # 디바이스 1개
        print(device)
        ondevice = device[0]
        print(ondevice)
        ondevice.shell("input keyevent KEYCODE_POWER")
        ondevice.shell("svc data disable")
        ondevice.shell("settings put global airplane_mode_on 1")
        ondevice.shell(
            "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state true")

        ondevice.shell("svc data enable")
        ondevice.shell("settings put global airplane_mode_on 0")
        ondevice.shell(
            "am broadcast -a android.intent.action.AIRPLANE_MODE --ez state false")
        print('아이피 변경 함??')
        time.sleep(3)
        while True:
            try:
                wait_float(0.5, 0.9)
                getIp = requests.get("https://api.ip.pe.kr/json/").json()['ip']
                if getIp is not None:
                    break
            except:
                continue
    except Exception as e:
        print(e)
        while True:
            try:
                wait_float(0.5, 0.9)
                getIp = requests.get("https://api.ip.pe.kr/json/").json()['ip']
                if getIp is not None:
                    break
            except:
                continue
    return getIp



def searchElement(ele,driver):
    wait_float(0.3, 0.7)
    re_count = 0
    element = ""
    while True:
        re_count += 1
        if re_count % 5 == 0:
            print(ele)
            print("새로고침!!!!")
            driver.refresh()
            
            pg.press('F5')
        elif element != "":
            break
        try:
            element = WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CSS_SELECTOR, ele)))
        except:
            pass
        

    selected_element = driver.find_elements(by=By.CSS_SELECTOR, value=ele)
    wait_float(0.3, 0.7)
    return selected_element




def wait_float(start, end):
    wait_ran = random.uniform(start, end)
    time.sleep(wait_ran)


def exitApp(driver):
    pg.alert(text='프로그램을 종료합니다.', title='제목입니다.', button='OK')
    try:
        driver.quit()
    except:
        pass
    sys.exit(0)


def focus_window(winName):
    while True:
        try:
            user32 = ctypes.windll.user32
            foreground_window = user32.GetForegroundWindow()
            window = gw.Window(foreground_window)
            if winName in window.title:
                break
            else:
                windows = Desktop(backend="uia").windows()
                for window in windows:
                    if winName in window.window_text():
                        window.set_focus()
                        break
        except Exception as e:
            print(str(e))
            pass


BASE_DIR = Path(__file__).resolve().parent






# subjectArr
def list_chunk(lst, n):
    return [lst[i:i+n] for i in range(0, len(lst), n)]


def isEnglishOrKorean(input_s):
    k_count = 0
    e_count = 0
    for c in input_s:
        if ord('가') <= ord(c) <= ord('힣'):
            k_count+=1
        elif ord('a') <= ord(c.lower()) <= ord('z'):
            e_count+=1
    return True if k_count>1 else False




def getNeighborFunc(getDict):
    
    if not getDict['neighbor_link']:
        pg.alert('링크가 없습니다. 확인 후 다시 실행해주세요')
        return

    getPostNum = getDict['neighbor_link'].split('/')[-1]

    neighborEx = load_workbook('./etc/neighbor_list.xlsx')
    nExSheet = neighborEx.active

    nExCount = 0
    while True:
        nExCount += 1
        if nExSheet.cell(nExCount,1).value is None:
            break
        
    # service = ChromeService(executable_path=ChromeDriverManager().install())
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service)
    
    while True:
        print('돈다돈다~')
        try:
            wait_float(0.5,1.2)
            driver.switch_to.frame('mainFrame')
            neighborListOpenBtn = driver.find_element(by=By.CSS_SELECTOR, value=f"#Sympathy{getPostNum}")
            neighborListOpenBtn.click()
        except:
            pass

        try:
            wait_float(0.5,1.2)
            neighborListBox = driver.find_element(by=By.CSS_SELECTOR, value=f".area_sympathy.pcol3")
            if neighborListBox:
                break
        except:
            pass


    
    while True:
        # driver.switch_to.default_content()
        try:
            wait_float(0.5,1.2)
            driver.switch_to.frame(f'sympathyFrm{getPostNum}')
            # driver.switch_to.frame(f'sympathyFrm{nowBlogLinkSplit[-1]}')
            wait_float(0.5,1.2)
            neighborListWrap = driver.find_element(by=By.CSS_SELECTOR, value='.wrap_blog2_sympathy')
            neighborList = neighborListWrap.find_elements(by=By.CSS_SELECTOR, value=".area_profile")
            if neighborList:
                break
        except:
            pass
    
    for idx, neighbor in enumerate(neighborList):
        idCount = nExCount + idx
        neighborLink = neighbor.find_element(by=By.CSS_SELECTOR, value='.link.pcol2').get_attribute('href')

        chkCount = 0
        overlapVal = False
        while True:
            chkCount += 1
            if nExSheet.cell(chkCount,1).value == neighborLink:
                overlapVal = True
                nExCount = nExCount - 1
                print('중복!')
                break
            elif nExSheet.cell(chkCount,1).value is None:
                break
        
        if not overlapVal:
            print('엑셀에 쓰기')
            nExSheet.cell(idCount,1).value = neighborLink
            neighborEx.save('./etc/neighbor_list.xlsx')
    pg.alert('종료합니당~')


def insert_img_line(contentArr):
    # 1. 각 요소의 앞뒤 공백 제거
    while True:
        try:
            contentArr = [line.strip() for line in contentArr]
            getContentArrLen = len(contentArr)
            insertIndex = random.randint(1, getContentArrLen - 1)
            contentArr.insert(insertIndex, 'img_line')
            return contentArr
        except Exception as e:
            print(str(e))
            return False

def insert_img_line_two(contentArr):
    """
    각 요소의 맨 앞과 맨 뒤의 공백을 제거하고,
    '다.'로 끝나는 요소 뒤에 랜덤으로 'img_line'을 삽입합니다.
    'img_line'은 연속되지 않으며 최소 3개 이상의 인덱스 차이를 유지합니다.

    :param contentArr: 원본 리스트
    :return: 'img_line'이 삽입된 새로운 리스트
    """
    # 1. 각 요소의 앞뒤 공백 제거
    contentArr = [line.strip() for line in contentArr]

    print('1. 각 요소의 앞뒤 공백 제거')

    # 2. '다.'로 끝나는 요소의 인덱스 찾기
    eligible_positions = [
        i for i, element in enumerate(contentArr) if element.endswith('다.')
    ]

    print('2. "다." 로 끝나는 요소의 인덱스 찾기')

    # 3. 랜덤으로 두 개의 위치 선택 (최소 3개의 인덱스 차이를 유지)
    selected_positions = set()

    errCount = 0
    while len(selected_positions) < 2 and len(eligible_positions) > 1:
        errCount += 1
        print(errCount)
        if errCount > 5:
            return False
        pos = random.choice(eligible_positions)
        if all(abs(pos - s) >= 1 for s in selected_positions):
            selected_positions.add(pos)
        wait_float(1.2,1.5)

    print('3. 랜덤으로 두 개의 위치 선택 (최소 3개의 인덱스 차이를 유지)')
    # 삽입을 위해 위치 정렬
    selected_positions = sorted(selected_positions)
    print('삽입을 위해 위치 정렬')

    # 4. 원본 리스트를 복사하여 새로운 리스트 생성
    resultArr = contentArr[:]

    print('4. 원본 리스트를 복사하여 새로운 리스트 생성')

    # 5. 'img_line' 삽입
    for i, pos in enumerate(selected_positions):
        resultArr.insert(pos + 1 + i, 'img_line')  # 요소 뒤에 삽입
    
    print('5. img_line 삽입')

    return resultArr


def split_elements_by_newline(contentArr):
    """
    주어진 배열의 요소에서 개행("\n")을 기준으로 요소를 나누어 배열을 다시 생성합니다.
    'img_line' 요소는 나누지 않고 유지합니다.

    :param contentArr: 원본 리스트
    :return: 개행으로 분리된 새 리스트
    """
    resultArr = []

    for element in contentArr:
        # 'img_line' 요소는 그대로 유지
        if element == 'img_line':
            resultArr.append(element)
        else:
            # 개행을 기준으로 나누고 각 부분을 리스트에 추가
            lines = element.split("\n")
            resultArr.extend(line.strip() for line in lines if line.strip())

    return resultArr


def replace_img_line_with_files(contentArr, file_list):
    """
    주어진 리스트에서 'img_line' 문자열을 'img_line|파일명'으로 대체합니다.

    :param contentArr: 원본 리스트
    :param file_list: 대체할 파일명 리스트
    :return: 대체된 새로운 리스트
    """
    resultArr = []
    file_index = 0

    for element in contentArr:
        if element == "img_line" and file_index < len(file_list):
            # 'img_line'을 'img_line|파일명'으로 대체
            resultArr.append(f"img_line|{file_list[file_index]}")
            file_index += 1
        else:
            # 'img_line'이 아니거나 파일명이 부족한 경우 원래 값을 유지
            resultArr.append(element)

    return resultArr


def remove_plain_img_line(contentArr):
    """
    리스트에서 'img_line'만 있는 요소를 제거합니다.

    :param contentArr: 원본 리스트
    :return: 'img_line'만 제거된 새로운 리스트
    """
    # 'img_line'만 있는 요소를 제외한 리스트 생성
    return [element for element in contentArr if element != 'img_line']


def has_img_line(lst):
    for item in lst:
        if 'img_line' in item:
            return True
    return False



def newsWorkFunc(driver, newsWorkCount):

    errCount = 0
    while True:
        errCount += 1
        if errCount > 5:
            focus_window('Chrome')
            pg.press('F5')
            errCount = 0
            wait_float(2.5,3.5)
            continue
        try:
            wait_float(0.3,0.5)
            mainMenuList = driver.find_elements(by=By.CSS_SELECTOR, value=".shs_item")
            for menu in mainMenuList:
                if '뉴스' in menu.text:
                    menu.click()
                    break
        except:
            pass

        

        try:
            wait_float(0.3,0.5)
            menuSelectedChk = driver.find_element(by=By.CSS_SELECTOR, value=".nav_link.nav_news")
            chkVal = menuSelectedChk.get_attribute('aria-selected')
            print(chkVal)
            if chkVal == 'true':
                break
        except:
            pass

        try:
            wait_float(1.2,1.9)
            closePopup = driver.find_element(by=By.CSS_SELECTOR, value=".lst_btn_close")
            print('새로운 팝업창 클릭 확인!')
            closePopup.click()
        except:
            pass

    
    print(newsWorkCount)
    newsClickNumList = []
    for i in range(newsWorkCount):
        errCount = 0
        errStatus = False
        while True:
            errCount += 1
            if errCount > 5:
                errStatus = True
                break
            print('뉴스 기사 하나 클릭!!')
            try:
                wait_float(2.5,3.5)
                getNewsList1 = driver.find_elements(by=By.CSS_SELECTOR, value=".cc_text_item")
                getNewsList2 = driver.find_elements(by=By.CSS_SELECTOR, value=".cnf_news_item")
                getNewsList = getNewsList1 + getNewsList2

                
                randomNewsVal = random.randint(0, len(getNewsList))
                print(randomNewsVal)
                if randomNewsVal in newsClickNumList:
                    continue
                getNewsList[randomNewsVal].click()
                newsClickNumList.append(randomNewsVal)
            except:
                pass

            try:
                wait_float(0.3,0.5)
                postInnerChk = driver.find_elements(by=By.CSS_SELECTOR, value=".ofhd_float_back")
                if len(postInnerChk) > 0:
                    break
            except:
                pass
        scrollRandomVal = random.randint(5,8)
        wait_float(2.5,3.5)
        for i in range(scrollRandomVal):
            print('스크롤 하기!!')
            pg.moveTo(300,500)
            scrollRangeVal = random.randint(350,450)
            pg.scroll(-scrollRangeVal)
            wait_float(3.5,4.5)

        
        while True:
            errCount += 1
            if errCount > 5:
                errCount = 0
                pg.press('F5')
                errStatus = True
            print('뒤로가기!!')
            try:
                wait_float(0.3,0.5)
                backToNewsBtn = driver.find_element(by=By.CSS_SELECTOR, value=".ofhd_float_back")
                backToNewsBtn.click()
            except:
                pass

            try:
                wait_float(0.3,0.5)
                menuSelectedChk = driver.find_elements(by=By.CSS_SELECTOR, value=".nav_link.nav_news")
                if len(menuSelectedChk) > 0:
                    break
            except:
                pass
            
        if errStatus == True:
            break
#>>>>>>>>>>>>>>>>>>>>>>> 안쓰는 함수



